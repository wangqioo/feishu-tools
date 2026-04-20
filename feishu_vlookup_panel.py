#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
飞书表格函数面板 v1.0
跨飞书表格的可视化函数操作工具
支持 VLOOKUP / XLOOKUP / INDEX-MATCH / SUMIF / COUNTIF / SUMIFS
"""

# ─── 依赖自动安装 ─────────────────────────────────────────────────────────────
import subprocess, sys, importlib

_DEPS = ["requests", "openpyxl"]

def _ensure_deps():
    for pkg in _DEPS:
        try:
            importlib.import_module(pkg)
        except ImportError:
            print(f"正在安装依赖: {pkg} ...")
            subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "-q"])

_ensure_deps()

# ─── 标准库 ───────────────────────────────────────────────────────────────────
import os, json, sqlite3, re, threading, time, copy
from datetime import datetime
from pathlib import Path

# ─── 第三方库 ─────────────────────────────────────────────────────────────────
import requests
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ─── Tkinter ─────────────────────────────────────────────────────────────────
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, scrolledtext

# ─── 常量 ─────────────────────────────────────────────────────────────────────
APP_NAME    = "飞书表格函数面板"
APP_VERSION = "1.0.0"
DATA_DIR    = Path("feishu_vlookup_data")
CONFIG_FILE = DATA_DIR / "config.json"
CACHE_DB    = DATA_DIR / "cache.db"

FUNCTION_DEFS = {
    "VLOOKUP":    "垂直查找  在查找列搜索关键字，返回同行指定列的值",
    "XLOOKUP":    "扩展查找  支持返回多列结果，可指定未找到时的默认值",
    "INDEX/MATCH":"索引匹配  最灵活的查找，支持横向/纵向/精确/模糊",
    "SUMIF":      "条件求和  统计满足条件的行对应列的数值之和",
    "COUNTIF":    "条件计数  统计满足条件的行数",
    "SUMIFS":     "多条件求和  同时满足多个条件时对数值列求和",
}

MATCH_MODES = [
    ("exact",      "精确匹配"),
    ("contains",   "包含匹配"),
    ("startswith", "前缀匹配"),
    ("endswith",   "后缀匹配"),
    ("regex",      "正则匹配"),
]

# ─── 颜色主题 ─────────────────────────────────────────────────────────────────
COLORS = {
    "bg":           "#F5F7FA",
    "card":         "#FFFFFF",
    "primary":      "#1664FF",
    "primary_dark": "#0D4FCC",
    "success":      "#00B578",
    "warning":      "#FF9F0A",
    "error":        "#FF3B30",
    "text":         "#1D2129",
    "text_sub":     "#86909C",
    "border":       "#E5E6EB",
    "row_even":     "#F7F8FA",
    "row_odd":      "#FFFFFF",
    "highlight":    "#E8F0FE",
    "tag_vlookup":  "#E8F4FD",
    "tag_sumif":    "#FFF7E6",
    "tag_index":    "#F0FFF4",
}

# ═════════════════════════════════════════════════════════════════════════════
# API 层
# ═════════════════════════════════════════════════════════════════════════════

class FeishuAPIClient:
    """飞书表格数据接口（兼容企业内网代理和官方 API）"""

    DEFAULT_PROXY  = "https://mcenter.huaqin.com"
    DEFAULT_APP_ID = "cli_a96ac38049f8d0e5"

    def __init__(self, base_url: str = None, app_id: str = None, user_id: str = None,
                 access_token: str = None):
        self.base_url     = (base_url or self.DEFAULT_PROXY).rstrip("/")
        self.app_id       = app_id    or self.DEFAULT_APP_ID
        self.user_id      = user_id   or ""
        self.access_token = access_token or ""

    # ── 内网代理模式 ──────────────────────────────────────────────────────────

    def _proxy_headers(self) -> dict:
        return {
            "Content-Type": "application/json",
            "Origin":       self.app_id,
            "userId":       self.user_id,
        }

    def get_sheet_meta_proxy(self, token: str) -> dict:
        url  = f"{self.base_url}/fs/sheet/v1/spreadsheetsMetainfo"
        body = {"spreadsheetToken": token}
        r = requests.post(url, json=body, headers=self._proxy_headers(), timeout=30)
        r.raise_for_status()
        return r.json()

    def get_sheet_values_proxy(self, token: str, sheet_id: str,
                               start_row: int = 1, end_row: int = 2000) -> dict:
        url  = f"{self.base_url}/fs/sheet/v1/getSheetsValue"
        body = {
            "spreadsheetToken": token,
            "sheetId":          sheet_id,
            "startRow":         start_row,
            "endRow":           end_row,
        }
        r = requests.post(url, json=body, headers=self._proxy_headers(), timeout=60)
        r.raise_for_status()
        return r.json()

    # ── 官方 Open API 模式 ────────────────────────────────────────────────────

    def _open_headers(self) -> dict:
        return {
            "Content-Type":  "application/json; charset=utf-8",
            "Authorization": f"Bearer {self.access_token}",
        }

    def get_sheet_meta_open(self, token: str) -> dict:
        url = f"https://open.feishu.cn/open-apis/sheets/v3/spreadsheets/{token}/sheets/query"
        r = requests.get(url, headers=self._open_headers(), timeout=30)
        r.raise_for_status()
        return r.json()

    def get_sheet_values_open(self, token: str, sheet_id: str,
                              range_str: str = None) -> dict:
        rng = range_str or f"{sheet_id}!A1:ZZ5000"
        url = (f"https://open.feishu.cn/open-apis/sheets/v3/spreadsheets"
               f"/{token}/values/{rng}")
        r = requests.get(url, headers=self._open_headers(), timeout=60)
        r.raise_for_status()
        return r.json()

    # ── 统一调用（根据配置自动选择） ──────────────────────────────────────────

    def fetch_meta(self, token: str, use_open_api: bool = False
                   ) -> tuple[list[dict], str]:
        """返回 (sheets, spreadsheet_title)
        sheets: [{sheet_id, sheet_title}, ...]
        spreadsheet_title: 整个表格的标题
        """
        if use_open_api:
            data     = self.get_sheet_meta_open(token) or {}
            inner    = data.get("data") or {}
            sheets   = inner.get("sheets") or []
            sp_title = inner.get("title") or ""
            return ([{"sheet_id": s["sheet_id"], "sheet_title": s["title"]} for s in sheets],
                    sp_title)
        else:
            data     = self.get_sheet_meta_proxy(token) or {}
            inner    = data.get("data") or {}
            sheets   = inner.get("sheets") or []
            sp_title = inner.get("title") or ""
            return ([{"sheet_id": s.get("sheetId", ""), "sheet_title": s.get("title", "")}
                     for s in sheets],
                    sp_title)

    def fetch_values(self, token: str, sheet_id: str,
                     use_open_api: bool = False,
                     max_rows: int = 5000) -> list[list]:
        """返回二维列表 rows[row][col]"""
        if use_open_api:
            data       = self.get_sheet_values_open(token, sheet_id) or {}
            inner      = data.get("data") or {}
            value_range = inner.get("valueRange") or {}
            return value_range.get("values") or []
        else:
            data  = self.get_sheet_values_proxy(token, sheet_id, 1, max_rows) or {}
            inner = data.get("data") or {}
            return inner.get("values") or []

    # ── 写入 ──────────────────────────────────────────────────────────────────

    @staticmethod
    def _col_letter(n: int) -> str:
        """0-based 列索引 → Excel 列字母 (0→A, 25→Z, 26→AA …)"""
        result = ""
        n += 1
        while n:
            n, rem = divmod(n - 1, 26)
            result = chr(65 + rem) + result
        return result

    def write_values_proxy(self, token: str, sheet_id: str,
                           values: list[list],
                           start_row: int = 0, start_col: int = 0) -> dict:
        """内网代理写入（行列均 0-based，0 行为表头行）"""
        if not values:
            return {}
        url  = f"{self.base_url}/fs/sheet/v1/setSheetsValue"
        body = {
            "spreadsheetToken": token,
            "sheetId":          sheet_id,
            "startRow":         start_row,
            "startCol":         start_col,
            "values":           values,
        }
        r = requests.post(url, json=body, headers=self._proxy_headers(), timeout=60)
        r.raise_for_status()
        return r.json()

    def write_values_open(self, token: str, sheet_id: str,
                          values: list[list],
                          start_row: int = 0, start_col: int = 0) -> dict:
        """飞书开放平台写入"""
        if not values:
            return {}
        end_row = start_row + len(values)
        end_col = start_col + max((len(r) for r in values), default=1)
        rng = (f"{sheet_id}!"
               f"{self._col_letter(start_col)}{start_row + 1}:"
               f"{self._col_letter(end_col - 1)}{end_row}")
        url  = (f"https://open.feishu.cn/open-apis/sheets/v3/spreadsheets"
                f"/{token}/values")
        body = {"valueRange": {"range": rng, "values": values}}
        r = requests.put(url, json=body, headers=self._open_headers(), timeout=60)
        r.raise_for_status()
        return r.json()

    def write_values(self, token: str, sheet_id: str,
                     values: list[list],
                     start_row: int = 0, start_col: int = 0,
                     use_open_api: bool = False) -> dict:
        """统一写入接口（start_row/start_col 均 0-based）"""
        if use_open_api:
            return self.write_values_open(token, sheet_id, values, start_row, start_col)
        else:
            return self.write_values_proxy(token, sheet_id, values, start_row, start_col)


# ═════════════════════════════════════════════════════════════════════════════
# 数据缓存层
# ═════════════════════════════════════════════════════════════════════════════

class DataCache:
    """SQLite 本地缓存，存储各飞书 Sheet 的数据快照"""

    def __init__(self, db_path: Path = CACHE_DB):
        DATA_DIR.mkdir(exist_ok=True)
        self.conn = sqlite3.connect(str(db_path), check_same_thread=False)
        self._init_tables()

    def _init_tables(self):
        c = self.conn.cursor()
        # 数据源配置表
        c.execute("""CREATE TABLE IF NOT EXISTS sources (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            name         TEXT NOT NULL,
            token        TEXT NOT NULL,
            sheet_id     TEXT NOT NULL,
            sheet_title  TEXT,
            use_open_api INTEGER DEFAULT 0,
            row_count    INTEGER DEFAULT 0,
            col_count    INTEGER DEFAULT 0,
            synced_at    TEXT,
            source_type  TEXT DEFAULT 'online',
            UNIQUE(token, sheet_id)
        )""")
        # 兼容旧数据库：若没有 source_type 列则补充
        try:
            c.execute("ALTER TABLE sources ADD COLUMN source_type TEXT DEFAULT 'online'")
        except sqlite3.OperationalError:
            pass  # 列已存在，忽略
        # 通用数据行表（按 source_id 分区）
        c.execute("""CREATE TABLE IF NOT EXISTS sheet_rows (
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            source_id INTEGER NOT NULL,
            row_idx   INTEGER NOT NULL,
            col_idx   INTEGER NOT NULL,
            value     TEXT,
            FOREIGN KEY(source_id) REFERENCES sources(id) ON DELETE CASCADE
        )""")
        c.execute("CREATE INDEX IF NOT EXISTS idx_rows_source ON sheet_rows(source_id, row_idx)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_rows_col    ON sheet_rows(source_id, col_idx, value)")
        # API 配置
        c.execute("""CREATE TABLE IF NOT EXISTS api_config (
            key   TEXT PRIMARY KEY,
            value TEXT
        )""")
        self.conn.commit()

    # ── 数据源管理 ────────────────────────────────────────────────────────────

    def add_source(self, name: str, token: str, sheet_id: str,
                   sheet_title: str = "", use_open_api: bool = False,
                   source_type: str = "online") -> int:
        c = self.conn.cursor()
        c.execute("""INSERT OR REPLACE INTO sources
                     (name, token, sheet_id, sheet_title, use_open_api, source_type)
                     VALUES (?,?,?,?,?,?)""",
                  (name, token, sheet_id, sheet_title, int(use_open_api), source_type))
        self.conn.commit()
        return c.lastrowid

    def list_sources(self) -> list[dict]:
        c = self.conn.cursor()
        c.execute("SELECT id,name,token,sheet_id,sheet_title,use_open_api,"
                  "row_count,col_count,synced_at,source_type FROM sources ORDER BY id")
        cols = [d[0] for d in c.description]
        return [dict(zip(cols, row)) for row in c.fetchall()]

    def delete_source(self, source_id: int):
        c = self.conn.cursor()
        c.execute("DELETE FROM sheet_rows WHERE source_id=?", (source_id,))
        c.execute("DELETE FROM sources WHERE id=?", (source_id,))
        self.conn.commit()

    def get_source(self, source_id: int) -> dict | None:
        c = self.conn.cursor()
        c.execute("SELECT id,name,token,sheet_id,sheet_title,use_open_api,"
                  "row_count,col_count,synced_at,source_type FROM sources WHERE id=?",
                  (source_id,))
        row = c.fetchone()
        if not row:
            return None
        cols = [d[0] for d in c.description]
        return dict(zip(cols, row))

    # ── 数据写入 / 读取 ───────────────────────────────────────────────────────

    def store_rows(self, source_id: int, rows: list[list]):
        c = self.conn.cursor()
        c.execute("DELETE FROM sheet_rows WHERE source_id=?", (source_id,))
        batch = []
        for ri, row in enumerate(rows):
            for ci, val in enumerate(row):
                if val is not None:
                    batch.append((source_id, ri, ci, str(val)))
        if batch:
            c.executemany("INSERT INTO sheet_rows(source_id,row_idx,col_idx,value)"
                          " VALUES(?,?,?,?)", batch)
        row_count = len(rows)
        col_count = max((len(r) for r in rows), default=0)
        c.execute("UPDATE sources SET row_count=?,col_count=?,synced_at=? WHERE id=?",
                  (row_count, col_count, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), source_id))
        self.conn.commit()

    def get_headers(self, source_id: int) -> list[str]:
        """返回第 0 行（表头）"""
        c = self.conn.cursor()
        c.execute("SELECT col_idx, value FROM sheet_rows "
                  "WHERE source_id=? AND row_idx=0 ORDER BY col_idx", (source_id,))
        rows = c.fetchall()
        if not rows:
            return []
        max_col = rows[-1][0]
        headers = [""] * (max_col + 1)
        for ci, val in rows:
            headers[ci] = val or ""
        return headers

    def get_all_rows(self, source_id: int) -> list[list[str]]:
        """返回所有行（含表头），每行是等长的字符串列表"""
        src = self.get_source(source_id)
        if not src:
            return []
        col_count = src["col_count"] or 1
        c = self.conn.cursor()
        c.execute("SELECT row_idx, col_idx, value FROM sheet_rows "
                  "WHERE source_id=? ORDER BY row_idx, col_idx", (source_id,))
        from collections import defaultdict
        row_map: dict[int, dict[int, str]] = defaultdict(dict)
        for ri, ci, val in c.fetchall():
            row_map[ri][ci] = val or ""
        if not row_map:
            return []
        max_row = max(row_map.keys())
        result = []
        for ri in range(max_row + 1):
            row = [row_map[ri].get(ci, "") for ci in range(col_count)]
            result.append(row)
        return result

    def get_col_values(self, source_id: int, col_idx: int,
                       skip_header: bool = True) -> list[str]:
        c = self.conn.cursor()
        start = 1 if skip_header else 0
        c.execute("SELECT value FROM sheet_rows WHERE source_id=? AND col_idx=?"
                  " AND row_idx>=? ORDER BY row_idx", (source_id, col_idx, start))
        return [r[0] or "" for r in c.fetchall()]

    # ── API 配置 ──────────────────────────────────────────────────────────────

    def save_api_config(self, cfg: dict):
        c = self.conn.cursor()
        for k, v in cfg.items():
            c.execute("INSERT OR REPLACE INTO api_config(key,value) VALUES(?,?)", (k, str(v)))
        self.conn.commit()

    def load_api_config(self) -> dict:
        c = self.conn.cursor()
        c.execute("SELECT key,value FROM api_config")
        return {k: v for k, v in c.fetchall()}


# ═════════════════════════════════════════════════════════════════════════════
# 公式引擎
# ═════════════════════════════════════════════════════════════════════════════

class FormulaEngine:
    """在内存二维表上执行类 Excel 函数"""

    @staticmethod
    def _match_value(cell: str, keyword: str, mode: str) -> bool:
        c, k = str(cell).strip(), str(keyword).strip()
        if mode == "exact":
            return c.upper() == k.upper()
        if mode == "contains":
            return k.upper() in c.upper()
        if mode == "startswith":
            return c.upper().startswith(k.upper())
        if mode == "endswith":
            return c.upper().endswith(k.upper())
        if mode == "regex":
            try:
                return bool(re.search(k, c, re.IGNORECASE))
            except re.error:
                return False
        return c.upper() == k.upper()

    # ── VLOOKUP ───────────────────────────────────────────────────────────────

    @staticmethod
    def vlookup(lookup_values: list[str],
                lookup_table: list[list[str]],
                key_col: int,
                return_cols: list[int],
                match_mode: str = "exact",
                not_found: str = "#N/A") -> list[dict]:
        """
        lookup_values: 查找键列表
        lookup_table:  来源表（list of rows，含表头）
        key_col:       查找列索引（0-based，在 lookup_table 中）
        return_cols:   要返回的列索引列表
        """
        # 构建索引加速精确匹配
        index: dict[str, list[str]] = {}
        for row in lookup_table[1:]:  # 跳过表头
            if key_col < len(row):
                k = str(row[key_col]).strip().upper()
                if k not in index:
                    index[k] = row

        results = []
        for lv in lookup_values:
            lv_str = str(lv).strip()
            found_row = None

            if match_mode == "exact":
                found_row = index.get(lv_str.upper())
            else:
                for row in lookup_table[1:]:
                    if key_col < len(row):
                        if FormulaEngine._match_value(row[key_col], lv_str, match_mode):
                            found_row = row
                            break

            if found_row is not None:
                returned = {
                    f"col_{rc}": (found_row[rc] if rc < len(found_row) else "")
                    for rc in return_cols
                }
                returned["__status__"] = "MATCH"
            else:
                returned = {f"col_{rc}": not_found for rc in return_cols}
                returned["__status__"] = "NOT_FOUND"

            returned["__lookup_value__"] = lv_str
            results.append(returned)
        return results

    # ── XLOOKUP ──────────────────────────────────────────────────────────────

    @staticmethod
    def xlookup(lookup_values: list[str],
                search_table: list[list[str]],
                search_col: int,
                return_table: list[list[str]],
                return_cols: list[int],
                match_mode: str = "exact",
                not_found: str = "#N/A") -> list[dict]:
        """
        支持 search_table 与 return_table 来自不同数据源（跨表查找）
        """
        # 建立行索引：search_table row_idx -> return_table row 映射
        results = []
        for lv in lookup_values:
            lv_str = str(lv).strip()
            found_ret_row = None

            for ri, row in enumerate(search_table[1:], start=1):
                if search_col < len(row):
                    if FormulaEngine._match_value(row[search_col], lv_str, match_mode):
                        if ri < len(return_table):
                            found_ret_row = return_table[ri]
                        break

            if found_ret_row is not None:
                returned = {
                    f"col_{rc}": (found_ret_row[rc] if rc < len(found_ret_row) else "")
                    for rc in return_cols
                }
                returned["__status__"] = "MATCH"
            else:
                returned = {f"col_{rc}": not_found for rc in return_cols}
                returned["__status__"] = "NOT_FOUND"

            returned["__lookup_value__"] = lv_str
            results.append(returned)
        return results

    # ── INDEX/MATCH ───────────────────────────────────────────────────────────

    @staticmethod
    def index_match(lookup_values: list[str],
                    match_table: list[list[str]],
                    match_col: int,
                    index_table: list[list[str]],
                    index_col: int,
                    match_mode: str = "exact",
                    not_found: str = "#N/A") -> list[dict]:
        """
        MATCH 在 match_table 的 match_col 列找到行号，
        INDEX 在 index_table 同行的 index_col 列取值
        支持跨表
        """
        results = []
        for lv in lookup_values:
            lv_str = str(lv).strip()
            matched_row_idx = None
            for ri, row in enumerate(match_table[1:], start=1):
                if match_col < len(row):
                    if FormulaEngine._match_value(row[match_col], lv_str, match_mode):
                        matched_row_idx = ri
                        break

            if matched_row_idx is not None and matched_row_idx < len(index_table):
                ret_row = index_table[matched_row_idx]
                val = ret_row[index_col] if index_col < len(ret_row) else ""
                results.append({
                    "__lookup_value__": lv_str,
                    "__status__":       "MATCH",
                    "result":           val,
                    "_row_idx":         matched_row_idx,
                })
            else:
                results.append({
                    "__lookup_value__": lv_str,
                    "__status__":       "NOT_FOUND",
                    "result":           not_found,
                    "_row_idx":         -1,
                })
        return results

    # ── SUMIF ─────────────────────────────────────────────────────────────────

    @staticmethod
    def sumif(table: list[list[str]],
              criteria_col: int,
              criteria: str,
              sum_col: int,
              match_mode: str = "exact") -> dict:
        total = 0.0
        count = 0
        matched_rows = []
        for row in table[1:]:
            if criteria_col < len(row):
                if FormulaEngine._match_value(row[criteria_col], criteria, match_mode):
                    raw = row[sum_col] if sum_col < len(row) else ""
                    try:
                        total += float(str(raw).replace(",", "").strip())
                        count += 1
                        matched_rows.append(row)
                    except ValueError:
                        pass
        return {"sum": total, "count": count, "matched_rows": matched_rows}

    # ── COUNTIF ───────────────────────────────────────────────────────────────

    @staticmethod
    def countif(table: list[list[str]],
                criteria_col: int,
                criteria: str,
                match_mode: str = "exact") -> dict:
        count = 0
        matched_rows = []
        for row in table[1:]:
            if criteria_col < len(row):
                if FormulaEngine._match_value(row[criteria_col], criteria, match_mode):
                    count += 1
                    matched_rows.append(row)
        return {"count": count, "matched_rows": matched_rows}

    # ── SUMIFS（多条件） ───────────────────────────────────────────────────────

    @staticmethod
    def sumifs(table: list[list[str]],
               sum_col: int,
               conditions: list[dict]) -> dict:
        """
        conditions: [{"col": int, "criteria": str, "mode": str}, ...]
        """
        total = 0.0
        count = 0
        matched_rows = []
        for row in table[1:]:
            all_match = all(
                (cond["col"] < len(row) and
                 FormulaEngine._match_value(row[cond["col"]], cond["criteria"], cond.get("mode", "exact")))
                for cond in conditions
            )
            if all_match:
                raw = row[sum_col] if sum_col < len(row) else ""
                try:
                    total += float(str(raw).replace(",", "").strip())
                    count += 1
                    matched_rows.append(row)
                except ValueError:
                    pass
        return {"sum": total, "count": count, "matched_rows": matched_rows}


# ═════════════════════════════════════════════════════════════════════════════
# Excel 导出
# ═════════════════════════════════════════════════════════════════════════════

class ExcelExporter:

    HEADER_FILL  = PatternFill("solid", fgColor="1664FF")
    HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
    MATCH_FILL   = PatternFill("solid", fgColor="E6F7F0")
    NOMATCH_FILL = PatternFill("solid", fgColor="FFF1F0")
    BORDER_SIDE  = Side(style="thin", color="E5E6EB")

    @classmethod
    def _border(cls):
        return Border(
            left=cls.BORDER_SIDE, right=cls.BORDER_SIDE,
            top=cls.BORDER_SIDE,  bottom=cls.BORDER_SIDE,
        )

    @classmethod
    def export_results(cls, path: str, headers: list[str],
                       rows: list[list[str]], status_col: int = None):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "函数结果"

        # 写表头
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill       = cls.HEADER_FILL
            cell.font       = cls.HEADER_FONT
            cell.alignment  = Alignment(horizontal="center", vertical="center")
            cell.border     = cls._border()
        ws.row_dimensions[1].height = 22

        # 写数据
        for ri, row in enumerate(rows, 2):
            fill = None
            if status_col is not None and status_col < len(row):
                fill = cls.MATCH_FILL if row[status_col] == "MATCH" else cls.NOMATCH_FILL
            for ci, val in enumerate(row, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.alignment = Alignment(vertical="center")
                cell.border    = cls._border()
                if fill:
                    cell.fill = fill

        # 自适应列宽
        for col in ws.columns:
            max_w = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_w + 4, 40)

        wb.save(path)


# ═════════════════════════════════════════════════════════════════════════════
# 本地文件读取器
# ═════════════════════════════════════════════════════════════════════════════

class LocalFileReader:
    """读取本地 xlsx / xls / csv 文件，返回二维字符串列表"""

    SUPPORTED = {
        ".xlsx": "Excel 2007+",
        ".xls":  "Excel 97-2003",
        ".csv":  "CSV 文本",
        ".tsv":  "TSV 文本",
    }

    @classmethod
    def can_read(cls, path: str) -> bool:
        return Path(path).suffix.lower() in cls.SUPPORTED

    @classmethod
    def read(cls, path: str, sheet_name: str = None) -> list[list[str]]:
        """
        返回二维字符串列表，第 0 行为表头。
        sheet_name: xlsx 时指定工作表名称，None 则取第一个工作表。
        """
        suffix = Path(path).suffix.lower()
        if suffix in (".xlsx", ".xls"):
            return cls._read_excel(path, sheet_name)
        elif suffix in (".csv", ".tsv"):
            return cls._read_csv(path, suffix)
        else:
            raise ValueError(f"不支持的文件格式: {suffix}")

    @classmethod
    def _read_excel(cls, path: str, sheet_name: str = None) -> list[list[str]]:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([("" if v is None else str(v)) for v in row])
        wb.close()
        # 去掉尾部全空行
        while rows and all(v == "" for v in rows[-1]):
            rows.pop()
        return rows

    @classmethod
    def _read_csv(cls, path: str, suffix: str) -> list[list[str]]:
        import csv
        delimiter = "\t" if suffix == ".tsv" else ","
        # 自动探测编码
        encodings = ["utf-8-sig", "gbk", "utf-8", "latin-1"]
        for enc in encodings:
            try:
                with open(path, "r", encoding=enc, newline="") as f:
                    reader = csv.reader(f, delimiter=delimiter)
                    rows = [list(row) for row in reader]
                while rows and all(v == "" for v in rows[-1]):
                    rows.pop()
                return rows
            except (UnicodeDecodeError, Exception):
                continue
        raise ValueError(f"无法解析文件编码: {path}")

    @classmethod
    def list_sheets(cls, path: str) -> list[str]:
        """列出 xlsx 文件的所有工作表名称"""
        suffix = Path(path).suffix.lower()
        if suffix not in (".xlsx", ".xls"):
            return []
        wb = openpyxl.load_workbook(path, read_only=True)
        names = list(wb.sheetnames)
        wb.close()
        return names


# ═════════════════════════════════════════════════════════════════════════════
# API 错误友好提示
# ═════════════════════════════════════════════════════════════════════════════

def _format_api_error(exc: Exception) -> tuple[str, str]:
    """
    将 API 异常转为 (title, detail) 用于 messagebox 展示。
    对 401/403 给出明确的权限指引。
    """
    import requests as _req

    msg = str(exc)

    # HTTP 状态码识别
    status = None
    if isinstance(exc, _req.HTTPError) and exc.response is not None:
        status = exc.response.status_code
        # 尝试解析响应体
        try:
            body = exc.response.json()
            code = body.get("code") or body.get("error_code") or ""
            server_msg = (body.get("msg") or body.get("message")
                          or body.get("error") or "")
            msg = f"[{code}] {server_msg}" if code else server_msg or msg
        except Exception:
            pass

    if status in (401, 403) or "401" in msg or "403" in msg or \
       "permission" in msg.lower() or "无权限" in msg or \
       "access denied" in msg.lower() or "not authorized" in msg.lower():
        title  = "权限不足，无法访问该飞书表格"
        detail = (
            f"错误信息: {msg}\n\n"
            "可能原因：\n"
            "  1. 当前账号没有该飞书表格的查看权限\n"
            "     → 请联系表格所有者，将你的账号添加为「可查看」\n\n"
            "  2. userId / App ID 填写有误\n"
            "     → 检查 API 设置页面的 userId 是否是你的工号\n\n"
            "  3. Access Token 已过期（使用官方 Open API 时）\n"
            "     → 重新生成 user_access_token 并填写到 API 设置页\n\n"
            "  4. 表格所在空间设置了组织外限制\n"
            "     → 联系空间管理员开放访问权限"
        )
        return title, detail

    if isinstance(exc, (_req.ConnectionError, _req.Timeout)):
        title  = "网络连接失败"
        detail = (
            f"错误信息: {msg}\n\n"
            "可能原因：\n"
            "  1. 当前不在公司内网，无法访问内网代理\n"
            "     → 连接 VPN 后重试，或切换至「飞书开放平台 API」模式\n\n"
            "  2. 代理地址配置错误\n"
            "     → 检查 API 设置中的「内网代理地址」"
        )
        return title, detail

    return "数据拉取失败", f"错误信息:\n{msg}"


# ═════════════════════════════════════════════════════════════════════════════
# GUI 组件 — 通用工具
# ═════════════════════════════════════════════════════════════════════════════

def _style_button(btn: ttk.Button, kind="primary"):
    """为按钮设置外观（ttk 主题无法完全控制，用 tk.Button 作为替代）"""
    pass  # 通过自定义 style 实现

def make_label_btn(parent, text, command, bg=None, fg="white",
                   padx=12, pady=4, font_size=9, radius=4):
    bg = bg or COLORS["primary"]
    btn = tk.Button(parent, text=text, command=command,
                    bg=bg, fg=fg, relief="flat", cursor="hand2",
                    padx=padx, pady=pady,
                    font=("Microsoft YaHei UI", font_size),
                    activebackground=COLORS["primary_dark"],
                    activeforeground="white",
                    bd=0)
    return btn

def make_card(parent, **kwargs):
    frame = tk.Frame(parent, bg=COLORS["card"],
                     relief="flat", bd=1,
                     highlightbackground=COLORS["border"],
                     highlightthickness=1,
                     **kwargs)
    return frame

def make_section_label(parent, text, sub=False):
    color = COLORS["text_sub"] if sub else COLORS["text"]
    size  = 8 if sub else 10
    lbl = tk.Label(parent, text=text, bg=COLORS["card"],
                   fg=color, font=("Microsoft YaHei UI", size, "bold" if not sub else ""))
    return lbl


# ═════════════════════════════════════════════════════════════════════════════
# GUI 面板 1 — 数据源管理
# ═════════════════════════════════════════════════════════════════════════════

class DataSourcePanel(tk.Frame):
    """管理飞书表格数据源（添加/刷新/删除/预览）"""

    def __init__(self, master, cache: DataCache, api: FeishuAPIClient,
                 on_sources_changed=None, **kwargs):
        super().__init__(master, bg=COLORS["bg"], **kwargs)
        self.cache   = cache
        self.api     = api
        self.on_change = on_sources_changed
        self._build_ui()
        self.refresh_table()

    def _build_ui(self):
        # ── 顶栏 ──────────────────────────────────────────────────────────────
        top = tk.Frame(self, bg=COLORS["bg"])
        top.pack(fill="x", padx=16, pady=(12, 6))

        tk.Label(top, text="数据源管理", bg=COLORS["bg"],
                 fg=COLORS["text"], font=("Microsoft YaHei UI", 13, "bold")).pack(side="left")

        btn_frame = tk.Frame(top, bg=COLORS["bg"])
        btn_frame.pack(side="right")
        make_label_btn(btn_frame, "+ 添加数据源", self._open_add_dialog).pack(side="left", padx=4)
        make_label_btn(btn_frame, "↻ 同步选中", self._sync_selected,
                       bg=COLORS["success"]).pack(side="left", padx=4)
        make_label_btn(btn_frame, "× 删除选中", self._delete_selected,
                       bg=COLORS["error"]).pack(side="left", padx=4)

        # ── 数据源列表 ────────────────────────────────────────────────────────
        list_card = make_card(self)
        list_card.pack(fill="both", expand=True, padx=16, pady=6)

        cols = ("id", "type", "name", "source", "sheet_title",
                "row_count", "col_count", "synced_at")
        self.tree = ttk.Treeview(list_card, columns=cols, show="headings",
                                 selectmode="extended", height=12)

        hdrs = {
            "id":          (40,  "ID"),
            "type":        (70,  "类型"),
            "name":        (150, "名称"),
            "source":      (200, "Token / 文件路径"),
            "sheet_title": (130, "工作表"),
            "row_count":   (60,  "行数"),
            "col_count":   (60,  "列数"),
            "synced_at":   (150, "同步时间"),
        }
        for col, (w, h) in hdrs.items():
            self.tree.heading(col, text=h)
            self.tree.column(col, width=w, minwidth=40)

        self.tree.tag_configure("online", foreground=COLORS["primary"])
        self.tree.tag_configure("local",  foreground=COLORS["success"])

        vsb = ttk.Scrollbar(list_card, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(list_card, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        list_card.rowconfigure(0, weight=1)
        list_card.columnconfigure(0, weight=1)

        self.tree.bind("<Double-1>", self._preview_source)

        # ── 预览区 ────────────────────────────────────────────────────────────
        prev_card = make_card(self)
        prev_card.pack(fill="both", expand=True, padx=16, pady=(4, 12))

        hdr = tk.Frame(prev_card, bg=COLORS["card"])
        hdr.pack(fill="x", padx=8, pady=4)
        tk.Label(hdr, text="数据预览（前 50 行）", bg=COLORS["card"],
                 fg=COLORS["text"], font=("Microsoft YaHei UI", 9, "bold")).pack(side="left")
        self.preview_info = tk.Label(hdr, text="", bg=COLORS["card"],
                                     fg=COLORS["text_sub"], font=("Microsoft YaHei UI", 8))
        self.preview_info.pack(side="right")

        self.preview_tree = ttk.Treeview(prev_card, show="headings",
                                         height=8, selectmode="none")
        psb = ttk.Scrollbar(prev_card, orient="vertical",
                            command=self.preview_tree.yview)
        phsb = ttk.Scrollbar(prev_card, orient="horizontal",
                             command=self.preview_tree.xview)
        self.preview_tree.configure(yscrollcommand=psb.set, xscrollcommand=phsb.set)
        self.preview_tree.pack(side="left", fill="both", expand=True, padx=(8, 0), pady=(0, 8))
        psb.pack(side="right", fill="y", pady=(0, 8))
        phsb.pack(side="bottom", fill="x", padx=8)

    # ── 刷新主列表 ────────────────────────────────────────────────────────────

    def refresh_table(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        for src in self.cache.list_sources():
            stype = src.get("source_type") or "online"
            type_label = "本地文件" if stype == "local" else "飞书在线"
            # 本地文件显示文件名，在线显示 token
            if stype == "local":
                source_display = Path(src["token"]).name
            else:
                source_display = src["token"]
            self.tree.insert("", "end", iid=str(src["id"]),
                             tags=(stype,),
                             values=(
                                 src["id"],
                                 type_label,
                                 src["name"],
                                 source_display,
                                 src["sheet_title"] or "-",
                                 src["row_count"] or "-",
                                 src["col_count"] or "-",
                                 src["synced_at"] or "未同步",
                             ))

    # ── 预览数据源 ────────────────────────────────────────────────────────────

    def _preview_source(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        source_id = int(sel[0])
        src = self.cache.get_source(source_id)
        if not src or not src["synced_at"]:
            messagebox.showinfo("提示", "该数据源尚未同步，请先同步数据", parent=self)
            return

        rows    = self.cache.get_all_rows(source_id)
        headers = rows[0] if rows else []
        data    = rows[1:51] if len(rows) > 1 else []

        # 重建预览表
        pt = self.preview_tree
        pt.delete(*pt.get_children())
        if not headers:
            return

        # 动态设置列
        col_ids = [f"c{i}" for i in range(len(headers))]
        pt["columns"] = col_ids
        for ci, (cid, h) in enumerate(zip(col_ids, headers)):
            pt.heading(cid, text=h or f"列{ci+1}")
            pt.column(cid, width=100, minwidth=60)

        for ri, row in enumerate(data):
            padded = row + [""] * (len(headers) - len(row))
            tag = "even" if ri % 2 == 0 else "odd"
            pt.insert("", "end", values=padded, tags=(tag,))

        pt.tag_configure("even", background=COLORS["row_even"])
        pt.tag_configure("odd",  background=COLORS["row_odd"])

        self.preview_info.config(
            text=f"共 {len(rows)-1} 行 × {len(headers)} 列 | 当前源: {src['name']}"
        )

    # ── 添加对话框 ────────────────────────────────────────────────────────────

    def _open_add_dialog(self):
        dlg = tk.Toplevel(self)
        dlg.title("添加数据源")
        dlg.geometry("600x640")
        dlg.resizable(False, False)
        dlg.grab_set()
        dlg.configure(bg=COLORS["bg"])

        tk.Label(dlg, text="添加数据源", bg=COLORS["bg"],
                 fg=COLORS["text"],
                 font=("Microsoft YaHei UI", 12, "bold")).pack(pady=(14, 2))
        tk.Label(dlg, text="支持在线飞书表格 和 本地 Excel/CSV 文件", bg=COLORS["bg"],
                 fg=COLORS["text_sub"],
                 font=("Microsoft YaHei UI", 9)).pack(pady=(0, 8))

        # ── 模式切换按钮（仿 Tab） ────────────────────────────────────────────
        mode_var = tk.StringVar(value="online")
        mode_bar = tk.Frame(dlg, bg=COLORS["bg"])
        mode_bar.pack(fill="x", padx=20, pady=(0, 4))

        online_btn = make_label_btn(mode_bar, "飞书在线表格",
                                    lambda: _switch_mode("online"),
                                    bg=COLORS["primary"])
        local_btn  = make_label_btn(mode_bar, "本地文件 (Excel/CSV)",
                                    lambda: _switch_mode("local"),
                                    bg=COLORS["border"], fg=COLORS["text"])
        online_btn.pack(side="left", padx=(0, 4))
        local_btn.pack(side="left")

        # 两个内容面板
        online_frame = make_card(dlg)
        local_frame  = make_card(dlg)
        for f in (online_frame, local_frame):
            f.pack(fill="x", padx=20, pady=4)
        local_frame.pack_forget()

        def _switch_mode(m: str):
            mode_var.set(m)
            if m == "online":
                online_frame.pack(fill="x", padx=20, pady=4, before=log_card)
                local_frame.pack_forget()
                online_btn.config(bg=COLORS["primary"], fg="white")
                local_btn.config(bg=COLORS["border"], fg=COLORS["text"])
            else:
                local_frame.pack(fill="x", padx=20, pady=4, before=log_card)
                online_frame.pack_forget()
                local_btn.config(bg=COLORS["success"], fg="white")
                online_btn.config(bg=COLORS["border"], fg=COLORS["text"])

        # ── 在线飞书表格参数 ──────────────────────────────────────────────────
        ol = {}  # online vars
        ol_fields = [
            ("数据源名称 *", "name",       "拉取工作表后自动填入表格标题，也可手动修改"),
            ("表格 Token *", "token",      "可直接粘贴飞书表格链接，自动提取 Token"),
            ("工作表 ID",    "sheet_id",   "留空后点「拉取工作表列表」自动填入"),
            ("工作表名称",   "sheet_title","可选，便于识别"),
        ]
        for ri, (label, key, hint) in enumerate(ol_fields):
            tk.Label(online_frame, text=label, bg=COLORS["card"],
                     fg=COLORS["text"],
                     font=("Microsoft YaHei UI", 9, "bold"),
                     width=14, anchor="e").grid(
                         row=ri*2, column=0, padx=(12, 4), pady=(8, 0))
            var = tk.StringVar()
            ol[key] = var
            ttk.Entry(online_frame, textvariable=var, width=44).grid(
                row=ri*2, column=1, padx=4, pady=(8, 0), sticky="ew")
            tk.Label(online_frame, text=hint, bg=COLORS["card"],
                     fg=COLORS["text_sub"],
                     font=("Microsoft YaHei UI", 7)).grid(
                         row=ri*2+1, column=1, padx=4, sticky="w")

        # Token 输入框：粘贴飞书链接时自动提取 token
        def _on_token_change(*_):
            val = ol["token"].get().strip()
            if "feishu.cn" in val or "larksuite.com" in val:
                import re
                m = re.search(
                    r'(?:feishu\.cn|larksuite\.com)/(?:sheets|base|wiki|docs|docx)/([A-Za-z0-9_-]+)',
                    val)
                if m:
                    ol["token"].set(m.group(1))
        ol["token"].trace_add("write", _on_token_change)

        # API 模式
        use_open = tk.BooleanVar(value=False)
        api_row = tk.Frame(online_frame, bg=COLORS["card"])
        api_row.grid(row=len(ol_fields)*2, column=0, columnspan=2,
                     padx=12, pady=(6, 0), sticky="w")
        tk.Label(api_row, text="API 模式:", bg=COLORS["card"],
                 fg=COLORS["text"],
                 font=("Microsoft YaHei UI", 9, "bold")).pack(side="left", padx=(0, 8))
        ttk.Radiobutton(api_row, text="内网代理 (默认)",
                        variable=use_open, value=False).pack(side="left", padx=4)
        ttk.Radiobutton(api_row, text="飞书开放平台 API",
                        variable=use_open, value=True).pack(side="left", padx=4)

        ol["max_rows"] = tk.StringVar(value="5000")
        mr_row = tk.Frame(online_frame, bg=COLORS["card"])
        mr_row.grid(row=len(ol_fields)*2+1, column=0, columnspan=2,
                    padx=12, pady=(0, 8), sticky="w")
        tk.Label(mr_row, text="最大拉取行数:",
                 bg=COLORS["card"], fg=COLORS["text"],
                 font=("Microsoft YaHei UI", 9, "bold")).pack(side="left")
        ttk.Entry(mr_row, textvariable=ol["max_rows"], width=8).pack(
            side="left", padx=8)
        tk.Label(mr_row, text="行（含表头）",
                 bg=COLORS["card"], fg=COLORS["text_sub"],
                 font=("Microsoft YaHei UI", 8)).pack(side="left")
        online_frame.columnconfigure(1, weight=1)

        # ── 本地文件参数 ──────────────────────────────────────────────────────
        lc = {}
        tk.Label(local_frame, text="数据源名称 *",
                 bg=COLORS["card"], fg=COLORS["text"],
                 font=("Microsoft YaHei UI", 9, "bold"),
                 width=14, anchor="e").grid(row=0, column=0, padx=(12, 4), pady=(10, 0))
        lc["name"] = tk.StringVar()
        ttk.Entry(local_frame, textvariable=lc["name"], width=42).grid(
            row=0, column=1, padx=4, pady=(10, 0), sticky="ew")

        tk.Label(local_frame, text="文件路径 *",
                 bg=COLORS["card"], fg=COLORS["text"],
                 font=("Microsoft YaHei UI", 9, "bold"),
                 width=14, anchor="e").grid(row=1, column=0, padx=(12, 4), pady=(8, 0))
        lc["filepath"] = tk.StringVar()
        path_row = tk.Frame(local_frame, bg=COLORS["card"])
        path_row.grid(row=1, column=1, padx=4, pady=(8, 0), sticky="ew")
        path_entry = ttk.Entry(path_row, textvariable=lc["filepath"], width=34)
        path_entry.pack(side="left")

        def _browse_file():
            fp = filedialog.askopenfilename(
                title="选择文件",
                filetypes=[
                    ("Excel 文件", "*.xlsx *.xls"),
                    ("CSV 文件", "*.csv *.tsv"),
                    ("所有支持的格式", "*.xlsx *.xls *.csv *.tsv"),
                ],
                parent=dlg,
            )
            if fp:
                lc["filepath"].set(fp)
                # 自动填充名称（如果还没填）
                if not lc["name"].get():
                    lc["name"].set(Path(fp).stem)
                # 如果是 xlsx，刷新工作表列表
                _refresh_sheets(fp)

        make_label_btn(path_row, "浏览...", _browse_file,
                       bg=COLORS["text_sub"], padx=8, pady=2,
                       font_size=8).pack(side="left", padx=6)

        tk.Label(local_frame, text="工作表",
                 bg=COLORS["card"], fg=COLORS["text"],
                 font=("Microsoft YaHei UI", 9, "bold"),
                 width=14, anchor="e").grid(row=2, column=0, padx=(12, 4), pady=(8, 0))
        lc["sheet"] = tk.StringVar()
        lc["sheet_cb"] = ttk.Combobox(local_frame, textvariable=lc["sheet"],
                                       width=30, state="readonly")
        lc["sheet_cb"].grid(row=2, column=1, padx=4, pady=(8, 0), sticky="w")
        tk.Label(local_frame, text="CSV/TSV 文件只有一个工作表，无需选择",
                 bg=COLORS["card"], fg=COLORS["text_sub"],
                 font=("Microsoft YaHei UI", 7)).grid(
                     row=3, column=1, padx=4, sticky="w", pady=(0, 8))
        local_frame.columnconfigure(1, weight=1)

        def _refresh_sheets(fp: str):
            sheets = LocalFileReader.list_sheets(fp)
            lc["sheet_cb"]["values"] = sheets
            if sheets:
                lc["sheet"].set(sheets[0])
            else:
                lc["sheet"].set("")

        lc["filepath"].trace_add("write", lambda *_: (
            _refresh_sheets(lc["filepath"].get())
            if lc["filepath"].get() else None
        ))

        # ── 状态日志 ─────────────────────────────────────────────────────────
        log_card = tk.Frame(dlg, bg=COLORS["bg"])
        log_card.pack(fill="x", padx=20, pady=4)
        log_var = tk.StringVar(value="等待操作...")
        tk.Label(log_card, textvariable=log_var, bg=COLORS["bg"],
                 fg=COLORS["text_sub"],
                 font=("Microsoft YaHei UI", 8), wraplength=540).pack(anchor="w")

        # ── 按钮 ─────────────────────────────────────────────────────────────
        btn_row = tk.Frame(dlg, bg=COLORS["bg"])
        btn_row.pack(pady=10)

        def do_fetch_sheets():
            token = ol["token"].get().strip()
            if not token:
                messagebox.showwarning("提示", "请填写表格 Token", parent=dlg)
                return
            log_var.set("正在拉取工作表列表...")
            dlg.update_idletasks()
            try:
                sheets, sp_title = self.api.fetch_meta(token, use_open_api=use_open.get())
                if not sheets:
                    log_var.set("未找到工作表，请检查 Token 或权限")
                    return
                ol["sheet_id"].set(sheets[0]["sheet_id"])
                ol["sheet_title"].set(sheets[0]["sheet_title"])
                # 自动填入数据源名称（表格标题 > 第一个工作表名，仅当名称栏为空时）
                if not ol["name"].get().strip():
                    auto_name = sp_title or sheets[0]["sheet_title"]
                    if auto_name:
                        ol["name"].set(auto_name)
                log_var.set(
                    f"发现 {len(sheets)} 个工作表，已填入第一个: "
                    f"{sheets[0]['sheet_title']}")
            except Exception as ex:
                title, detail = _format_api_error(ex)
                log_var.set(f"失败: {title}")
                messagebox.showerror(title, detail, parent=dlg)

        def do_add_online():
            name  = ol["name"].get().strip()
            token = ol["token"].get().strip()
            sid   = ol["sheet_id"].get().strip()
            title = ol["sheet_title"].get().strip()
            if not name or not token or not sid:
                messagebox.showwarning("提示", "名称、Token、工作表 ID 为必填项", parent=dlg)
                return
            try:
                max_r = int(ol["max_rows"].get().strip() or "5000")
            except ValueError:
                max_r = 5000
            log_var.set("正在拉取数据...")
            dlg.update_idletasks()

            def worker():
                try:
                    rows = self.api.fetch_values(token, sid,
                                                 use_open_api=use_open.get(),
                                                 max_rows=max_r)
                    src_id = self.cache.add_source(
                        name, token, sid, title,
                        use_open_api=use_open.get(),
                        source_type="online",
                    )
                    self.cache.store_rows(src_id, rows)
                    n = len(rows)
                    def _ok_online(n=n):
                        if not dlg.winfo_exists(): return
                        log_var.set(f"完成！共拉取 {n} 行数据")
                        self.refresh_table()
                        if self.on_change: self.on_change()
                    dlg.after(0, _ok_online)
                except Exception as ex:
                    err_title, err_detail = _format_api_error(ex)
                    def _err_online(t=err_title, d=err_detail):
                        if not dlg.winfo_exists(): return
                        log_var.set(f"失败: {t}")
                        messagebox.showerror(t, d, parent=dlg)
                    dlg.after(0, _err_online)

            threading.Thread(target=worker, daemon=True).start()

        def do_add_local():
            name = lc["name"].get().strip()
            fp   = lc["filepath"].get().strip()
            if not name or not fp:
                messagebox.showwarning("提示", "名称和文件路径为必填项", parent=dlg)
                return
            if not Path(fp).exists():
                messagebox.showerror("错误", f"文件不存在:\n{fp}", parent=dlg)
                return
            if not LocalFileReader.can_read(fp):
                messagebox.showerror("错误",
                                     f"不支持的文件格式: {Path(fp).suffix}\n"
                                     f"支持: {', '.join(LocalFileReader.SUPPORTED)}",
                                     parent=dlg)
                return
            sheet = lc["sheet"].get().strip() or None
            log_var.set("正在读取文件...")
            dlg.update_idletasks()

            def worker():
                try:
                    rows = LocalFileReader.read(fp, sheet_name=sheet)
                    title = sheet or Path(fp).stem
                    src_id = self.cache.add_source(
                        name,
                        token=fp,           # 用文件路径当 token
                        sheet_id=sheet or "sheet1",
                        sheet_title=title,
                        use_open_api=False,
                        source_type="local",
                    )
                    self.cache.store_rows(src_id, rows)
                    n = len(rows)
                    def _ok_local(n=n):
                        if not dlg.winfo_exists(): return
                        log_var.set(f"完成！共读取 {n} 行数据")
                        self.refresh_table()
                        if self.on_change: self.on_change()
                    dlg.after(0, _ok_local)
                except Exception as ex:
                    def _err_local(e=ex):
                        if not dlg.winfo_exists(): return
                        log_var.set(f"读取失败: {e}")
                        messagebox.showerror("读取失败", str(e), parent=dlg)
                    dlg.after(0, _err_local)

            threading.Thread(target=worker, daemon=True).start()

        def do_add():
            if mode_var.get() == "online":
                do_add_online()
            else:
                do_add_local()

        self._fetch_btn = make_label_btn(btn_row, "拉取工作表列表", do_fetch_sheets,
                                          bg=COLORS["warning"])
        self._fetch_btn.pack(side="left", padx=4)
        make_label_btn(btn_row, "添加并导入", do_add).pack(side="left", padx=4)
        make_label_btn(btn_row, "取消", dlg.destroy,
                       bg=COLORS["text_sub"]).pack(side="left", padx=4)

        def _update_buttons(*_):
            """在线模式显示「拉取工作表列表」，本地模式隐藏"""
            if mode_var.get() == "online":
                self._fetch_btn.pack(side="left", padx=4, before=btn_row.winfo_children()[1])
            else:
                self._fetch_btn.pack_forget()

        mode_var.trace_add("write", _update_buttons)

    # ── 同步 ─────────────────────────────────────────────────────────────────

    def _sync_selected(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("提示", "请先选中要同步的数据源", parent=self)
            return
        for iid in sel:
            src = self.cache.get_source(int(iid))
            if not src:
                continue

            stype = src.get("source_type") or "online"

            if stype == "local":
                # 本地文件：重新从磁盘读取
                def local_worker(s=src):
                    fp    = s["token"]   # 存文件路径
                    sheet = s["sheet_id"] if s["sheet_id"] != "sheet1" else None
                    try:
                        if not Path(fp).exists():
                            raise FileNotFoundError(
                                f"文件已不存在，请重新添加数据源:\n{fp}")
                        rows = LocalFileReader.read(fp, sheet_name=sheet)
                        self.cache.store_rows(s["id"], rows)
                        self.after(0, self.refresh_table)
                        self.after(0, lambda: self.on_change and self.on_change())
                    except Exception as ex:
                        self.after(0, lambda e=ex: messagebox.showerror(
                            "本地文件重新读取失败", str(e), parent=self))
                threading.Thread(target=local_worker, daemon=True).start()
            else:
                # 在线飞书：通过 API 拉取，详细处理权限错误
                def online_worker(s=src):
                    try:
                        rows = self.api.fetch_values(
                            s["token"], s["sheet_id"],
                            use_open_api=bool(s["use_open_api"])
                        )
                        self.cache.store_rows(s["id"], rows)
                        self.after(0, self.refresh_table)
                        self.after(0, lambda: self.on_change and self.on_change())
                    except Exception as ex:
                        err_title, err_detail = _format_api_error(ex)
                        self.after(0, lambda t=err_title, d=err_detail: messagebox.showerror(
                            t, d, parent=self))
                threading.Thread(target=online_worker, daemon=True).start()

    # ── 删除 ─────────────────────────────────────────────────────────────────

    def _delete_selected(self):
        sel = self.tree.selection()
        if not sel:
            return
        if not messagebox.askyesno("确认", f"删除选中的 {len(sel)} 个数据源？", parent=self):
            return
        for iid in sel:
            self.cache.delete_source(int(iid))
        self.refresh_table()
        if self.on_change:
            self.on_change()


# ═════════════════════════════════════════════════════════════════════════════
# GUI 面板 2 — 函数构建器
# ═════════════════════════════════════════════════════════════════════════════

class FunctionBuilderPanel(tk.Frame):
    """可视化函数参数配置 + 执行"""

    def __init__(self, master, cache: DataCache, on_results=None, **kwargs):
        super().__init__(master, bg=COLORS["bg"], **kwargs)
        self.cache      = cache
        self.on_results = on_results  # callback(headers, rows, func_type)
        self._sources: list[dict] = []
        self._rows_cache: dict[int, list[list[str]]] = {}
        self._build_ui()
        self.reload_sources()

    def reload_sources(self):
        self._sources = self.cache.list_sources()
        self._rows_cache.clear()
        self._update_source_dropdowns()

    def _get_rows(self, source_id: int) -> list[list[str]]:
        if source_id not in self._rows_cache:
            self._rows_cache[source_id] = self.cache.get_all_rows(source_id)
        return self._rows_cache[source_id]

    def _source_names(self) -> list[str]:
        return [f"{s['id']}: {s['name']} [{s['sheet_title'] or s['sheet_id']}]"
                for s in self._sources]

    def _source_by_choice(self, choice: str) -> dict | None:
        try:
            sid = int(choice.split(":")[0])
            return next((s for s in self._sources if s["id"] == sid), None)
        except (ValueError, IndexError):
            return None

    def _headers_for_source(self, choice: str) -> list[str]:
        src = self._source_by_choice(choice)
        if not src:
            return []
        rows = self._get_rows(src["id"])
        return rows[0] if rows else []

    def _build_ui(self):
        # ── 顶栏 ──────────────────────────────────────────────────────────────
        top = tk.Frame(self, bg=COLORS["bg"])
        top.pack(fill="x", padx=16, pady=(12, 4))
        tk.Label(top, text="函数构建器", bg=COLORS["bg"],
                 fg=COLORS["text"], font=("Microsoft YaHei UI", 13, "bold")).pack(side="left")

        # ── 函数类型选择 ──────────────────────────────────────────────────────
        type_card = make_card(self)
        type_card.pack(fill="x", padx=16, pady=4)
        tk.Label(type_card, text="选择函数类型", bg=COLORS["card"],
                 fg=COLORS["text"], font=("Microsoft YaHei UI", 9, "bold")).pack(
                     anchor="w", padx=12, pady=(8, 4))

        self._func_var = tk.StringVar(value="VLOOKUP")
        btn_row = tk.Frame(type_card, bg=COLORS["card"])
        btn_row.pack(fill="x", padx=12, pady=(0, 8))

        self._func_btns: dict[str, tk.Label] = {}
        for fname in FUNCTION_DEFS:
            lbl = tk.Label(btn_row, text=fname, bg=COLORS["border"],
                           fg=COLORS["text"], cursor="hand2",
                           font=("Microsoft YaHei UI", 9),
                           padx=12, pady=5, relief="flat")
            lbl.pack(side="left", padx=4)
            lbl.bind("<Button-1>", lambda e, f=fname: self._select_func(f))
            self._func_btns[fname] = lbl

        self.func_desc = tk.Label(type_card, text=FUNCTION_DEFS["VLOOKUP"],
                                  bg=COLORS["card"], fg=COLORS["text_sub"],
                                  font=("Microsoft YaHei UI", 8))
        self.func_desc.pack(anchor="w", padx=12, pady=(0, 8))

        # ── 参数面板容器（动态切换）────────────────────────────────────────────
        self._param_container = tk.Frame(self, bg=COLORS["bg"])
        self._param_container.pack(fill="both", expand=True, padx=16, pady=4)

        self._param_panels: dict[str, tk.Frame] = {}
        self._build_vlookup_panel()
        self._build_xlookup_panel()
        self._build_index_match_panel()
        self._build_sumif_panel()
        self._build_countif_panel()
        self._build_sumifs_panel()

        # ── 执行按钮栏 ────────────────────────────────────────────────────────
        run_bar = tk.Frame(self, bg=COLORS["bg"])
        run_bar.pack(fill="x", padx=16, pady=8)
        make_label_btn(run_bar, "▶  执行函数", self._run_function,
                       padx=20, pady=6, font_size=10).pack(side="left", padx=4)
        make_label_btn(run_bar, "↻ 重置参数", self._reset_params,
                       bg=COLORS["text_sub"]).pack(side="left", padx=4)

        self._run_info = tk.Label(run_bar, text="", bg=COLORS["bg"],
                                  fg=COLORS["text_sub"], font=("Microsoft YaHei UI", 8))
        self._run_info.pack(side="right", padx=8)

        self._select_func("VLOOKUP")

    def _select_func(self, fname: str):
        self._func_var.set(fname)
        for f, btn in self._func_btns.items():
            if f == fname:
                btn.config(bg=COLORS["primary"], fg="white")
            else:
                btn.config(bg=COLORS["border"], fg=COLORS["text"])
        self.func_desc.config(text=FUNCTION_DEFS.get(fname, ""))
        for f, panel in self._param_panels.items():
            if f == fname:
                panel.pack(fill="both", expand=True)
            else:
                panel.pack_forget()

    def _update_source_dropdowns(self):
        names = self._source_names()
        # 更新所有 Combobox（收集所有子组件中的 ttk.Combobox）
        for widget in self._param_container.winfo_children():
            self._update_cb_in_frame(widget, names)

    def _update_cb_in_frame(self, frame, names):
        for child in frame.winfo_children():
            if isinstance(child, ttk.Combobox) and getattr(child, "_is_source_cb", False):
                child["values"] = names
            self._update_cb_in_frame(child, names)

    # ── 构建各函数参数面板 ─────────────────────────────────────────────────────

    def _make_source_row(self, parent, row_i: int, label: str,
                         src_var: tk.StringVar, col_var: tk.StringVar,
                         col_label: str = "列") -> ttk.Combobox:
        """通用辅助：数据源 + 列选择行"""
        tk.Label(parent, text=label, bg=COLORS["card"], fg=COLORS["text"],
                 font=("Microsoft YaHei UI", 9, "bold"),
                 width=14, anchor="e").grid(row=row_i, column=0, padx=(12, 4), pady=6)

        src_cb = ttk.Combobox(parent, textvariable=src_var, width=34, state="readonly")
        src_cb["values"] = self._source_names()
        src_cb._is_source_cb = True
        src_cb.grid(row=row_i, column=1, padx=4, pady=6, sticky="w")

        tk.Label(parent, text=col_label, bg=COLORS["card"], fg=COLORS["text"],
                 font=("Microsoft YaHei UI", 9)).grid(row=row_i, column=2, padx=(8, 4))

        col_cb = ttk.Combobox(parent, textvariable=col_var, width=18, state="readonly")
        col_cb.grid(row=row_i, column=3, padx=4, pady=6, sticky="w")

        def on_src_change(*_):
            headers = self._headers_for_source(src_var.get())
            col_cb["values"] = [f"{i}: {h}" for i, h in enumerate(headers)]
            if headers:
                col_var.set(col_cb["values"][0])

        src_cb.bind("<<ComboboxSelected>>", on_src_change)
        return src_cb

    def _col_idx_from_var(self, var: tk.StringVar) -> int:
        v = var.get()
        try:
            return int(v.split(":")[0])
        except (ValueError, IndexError):
            return 0

    def _build_vlookup_panel(self):
        panel = make_card(self._param_container)
        self._param_panels["VLOOKUP"] = panel

        tk.Label(panel, text="VLOOKUP 参数", bg=COLORS["card"],
                 fg=COLORS["text"], font=("Microsoft YaHei UI", 10, "bold")).grid(
                     row=0, column=0, columnspan=4, padx=12, pady=(10, 4), sticky="w")

        self._vl = {}  # param vars

        self._vl["lv_src"]     = tk.StringVar()
        self._vl["lv_col"]     = tk.StringVar()
        self._vl["tbl_src"]    = tk.StringVar()
        self._vl["key_col"]    = tk.StringVar()
        self._vl["match_mode"] = tk.StringVar(value="exact")
        self._vl["not_found"]  = tk.StringVar(value="#N/A")

        self._make_source_row(panel, 1, "查找值来源", self._vl["lv_src"],
                              self._vl["lv_col"], "查找值列")
        self._make_source_row(panel, 2, "查找范围表格", self._vl["tbl_src"],
                              self._vl["key_col"], "查找键列")

        # 返回列（多选）
        tk.Label(panel, text="返回列（可多选）", bg=COLORS["card"], fg=COLORS["text"],
                 font=("Microsoft YaHei UI", 9, "bold"),
                 width=14, anchor="e").grid(row=3, column=0, padx=(12, 4), pady=6)
        self._vl_ret_frame = tk.Frame(panel, bg=COLORS["card"])
        self._vl_ret_frame.grid(row=3, column=1, columnspan=3, sticky="w", padx=4)
        self._vl["ret_lb"] = tk.Listbox(self._vl_ret_frame, selectmode="multiple",
                                         height=4, width=40)
        ret_sb = ttk.Scrollbar(self._vl_ret_frame, command=self._vl["ret_lb"].yview)
        self._vl["ret_lb"].configure(yscrollcommand=ret_sb.set)
        self._vl["ret_lb"].pack(side="left")
        ret_sb.pack(side="left", fill="y")

        def update_ret_lb(*_):
            headers = self._headers_for_source(self._vl["tbl_src"].get())
            lb = self._vl["ret_lb"]
            lb.delete(0, "end")
            for i, h in enumerate(headers):
                lb.insert("end", f"{i}: {h}")
        self._vl["tbl_src"].trace_add("write", update_ret_lb)

        # 匹配模式
        tk.Label(panel, text="匹配模式", bg=COLORS["card"], fg=COLORS["text"],
                 font=("Microsoft YaHei UI", 9, "bold"),
                 width=14, anchor="e").grid(row=4, column=0, padx=(12, 4), pady=6)
        mode_frame = tk.Frame(panel, bg=COLORS["card"])
        mode_frame.grid(row=4, column=1, columnspan=3, sticky="w", padx=4)
        for mode_val, mode_lbl in MATCH_MODES:
            ttk.Radiobutton(mode_frame, text=mode_lbl,
                            variable=self._vl["match_mode"],
                            value=mode_val).pack(side="left", padx=6)

        # 未找到时默认值
        tk.Label(panel, text="未找到默认值", bg=COLORS["card"], fg=COLORS["text"],
                 font=("Microsoft YaHei UI", 9, "bold"),
                 width=14, anchor="e").grid(row=5, column=0, padx=(12, 4), pady=6)
        ttk.Entry(panel, textvariable=self._vl["not_found"], width=20).grid(
            row=5, column=1, padx=4, sticky="w")

        panel.columnconfigure(1, weight=1)

    def _build_xlookup_panel(self):
        panel = make_card(self._param_container)
        self._param_panels["XLOOKUP"] = panel

        tk.Label(panel, text="XLOOKUP 参数（支持跨表查找）", bg=COLORS["card"],
                 fg=COLORS["text"], font=("Microsoft YaHei UI", 10, "bold")).grid(
                     row=0, column=0, columnspan=4, padx=12, pady=(10, 4), sticky="w")

        self._xl = {}
        self._xl["lv_src"]      = tk.StringVar()
        self._xl["lv_col"]      = tk.StringVar()
        self._xl["search_src"]  = tk.StringVar()
        self._xl["search_col"]  = tk.StringVar()
        self._xl["return_src"]  = tk.StringVar()
        self._xl["match_mode"]  = tk.StringVar(value="exact")
        self._xl["not_found"]   = tk.StringVar(value="#N/A")

        self._make_source_row(panel, 1, "查找值来源", self._xl["lv_src"],
                              self._xl["lv_col"], "查找值列")
        self._make_source_row(panel, 2, "搜索列来源", self._xl["search_src"],
                              self._xl["search_col"], "搜索列")

        # 返回数据源（可以不同于搜索源）
        tk.Label(panel, text="返回数据来源", bg=COLORS["card"], fg=COLORS["text"],
                 font=("Microsoft YaHei UI", 9, "bold"),
                 width=14, anchor="e").grid(row=3, column=0, padx=(12, 4), pady=6)
        ret_src_cb = ttk.Combobox(panel, textvariable=self._xl["return_src"],
                                   width=34, state="readonly")
        ret_src_cb["values"]    = self._source_names()
        ret_src_cb._is_source_cb = True
        ret_src_cb.grid(row=3, column=1, padx=4, sticky="w")

        tk.Label(panel, text="返回列（多选）", bg=COLORS["card"], fg=COLORS["text"],
                 font=("Microsoft YaHei UI", 9, "bold"),
                 width=14, anchor="e").grid(row=4, column=0, padx=(12, 4), pady=6)
        xl_ret_frame = tk.Frame(panel, bg=COLORS["card"])
        xl_ret_frame.grid(row=4, column=1, columnspan=3, sticky="w", padx=4)
        self._xl["ret_lb"] = tk.Listbox(xl_ret_frame, selectmode="multiple",
                                         height=4, width=40)
        xl_ret_sb = ttk.Scrollbar(xl_ret_frame, command=self._xl["ret_lb"].yview)
        self._xl["ret_lb"].configure(yscrollcommand=xl_ret_sb.set)
        self._xl["ret_lb"].pack(side="left")
        xl_ret_sb.pack(side="left", fill="y")

        def update_xl_ret(*_):
            headers = self._headers_for_source(self._xl["return_src"].get())
            lb = self._xl["ret_lb"]
            lb.delete(0, "end")
            for i, h in enumerate(headers):
                lb.insert("end", f"{i}: {h}")
        self._xl["return_src"].trace_add("write", update_xl_ret)

        tk.Label(panel, text="匹配模式", bg=COLORS["card"], fg=COLORS["text"],
                 font=("Microsoft YaHei UI", 9, "bold"),
                 width=14, anchor="e").grid(row=5, column=0, padx=(12, 4), pady=6)
        mode_frame = tk.Frame(panel, bg=COLORS["card"])
        mode_frame.grid(row=5, column=1, columnspan=3, sticky="w", padx=4)
        for mode_val, mode_lbl in MATCH_MODES:
            ttk.Radiobutton(mode_frame, text=mode_lbl,
                            variable=self._xl["match_mode"],
                            value=mode_val).pack(side="left", padx=6)

        tk.Label(panel, text="未找到默认值", bg=COLORS["card"], fg=COLORS["text"],
                 font=("Microsoft YaHei UI", 9, "bold"),
                 width=14, anchor="e").grid(row=6, column=0, padx=(12, 4), pady=6)
        ttk.Entry(panel, textvariable=self._xl["not_found"], width=20).grid(
            row=6, column=1, padx=4, sticky="w")

        panel.columnconfigure(1, weight=1)

    def _build_index_match_panel(self):
        panel = make_card(self._param_container)
        self._param_panels["INDEX/MATCH"] = panel

        tk.Label(panel, text="INDEX/MATCH 参数（支持跨表）", bg=COLORS["card"],
                 fg=COLORS["text"], font=("Microsoft YaHei UI", 10, "bold")).grid(
                     row=0, column=0, columnspan=4, padx=12, pady=(10, 4), sticky="w")

        self._im = {}
        self._im["lv_src"]      = tk.StringVar()
        self._im["lv_col"]      = tk.StringVar()
        self._im["match_src"]   = tk.StringVar()
        self._im["match_col"]   = tk.StringVar()
        self._im["index_src"]   = tk.StringVar()
        self._im["index_col"]   = tk.StringVar()
        self._im["match_mode"]  = tk.StringVar(value="exact")
        self._im["not_found"]   = tk.StringVar(value="#N/A")

        self._make_source_row(panel, 1, "查找值来源",  self._im["lv_src"],  self._im["lv_col"],  "查找值列")
        self._make_source_row(panel, 2, "MATCH 搜索表", self._im["match_src"], self._im["match_col"], "搜索列")
        self._make_source_row(panel, 3, "INDEX 返回表", self._im["index_src"], self._im["index_col"], "返回列")

        tk.Label(panel, text="匹配模式", bg=COLORS["card"], fg=COLORS["text"],
                 font=("Microsoft YaHei UI", 9, "bold"),
                 width=14, anchor="e").grid(row=4, column=0, padx=(12, 4), pady=6)
        mode_frame = tk.Frame(panel, bg=COLORS["card"])
        mode_frame.grid(row=4, column=1, columnspan=3, sticky="w", padx=4)
        for mode_val, mode_lbl in MATCH_MODES:
            ttk.Radiobutton(mode_frame, text=mode_lbl,
                            variable=self._im["match_mode"],
                            value=mode_val).pack(side="left", padx=6)

        tk.Label(panel, text="未找到默认值", bg=COLORS["card"], fg=COLORS["text"],
                 font=("Microsoft YaHei UI", 9, "bold"),
                 width=14, anchor="e").grid(row=5, column=0, padx=(12, 4), pady=6)
        ttk.Entry(panel, textvariable=self._im["not_found"], width=20).grid(
            row=5, column=1, padx=4, sticky="w")
        panel.columnconfigure(1, weight=1)

    def _build_sumif_panel(self):
        panel = make_card(self._param_container)
        self._param_panels["SUMIF"] = panel

        tk.Label(panel, text="SUMIF 参数", bg=COLORS["card"],
                 fg=COLORS["text"], font=("Microsoft YaHei UI", 10, "bold")).grid(
                     row=0, column=0, columnspan=4, padx=12, pady=(10, 4), sticky="w")

        self._si = {}
        self._si["tbl_src"]      = tk.StringVar()
        self._si["criteria_col"] = tk.StringVar()
        self._si["criteria"]     = tk.StringVar()
        self._si["sum_col"]      = tk.StringVar()
        self._si["match_mode"]   = tk.StringVar(value="exact")

        self._make_source_row(panel, 1, "数据表格", self._si["tbl_src"],
                              self._si["criteria_col"], "条件列")

        tk.Label(panel, text="条件值 *", bg=COLORS["card"], fg=COLORS["text"],
                 font=("Microsoft YaHei UI", 9, "bold"),
                 width=14, anchor="e").grid(row=2, column=0, padx=(12, 4), pady=6)
        ttk.Entry(panel, textvariable=self._si["criteria"], width=30).grid(
            row=2, column=1, padx=4, sticky="w")

        tk.Label(panel, text="求和列", bg=COLORS["card"], fg=COLORS["text"],
                 font=("Microsoft YaHei UI", 9, "bold"),
                 width=14, anchor="e").grid(row=3, column=0, padx=(12, 4), pady=6)
        self._si_sum_cb = ttk.Combobox(panel, textvariable=self._si["sum_col"],
                                        width=20, state="readonly")
        self._si_sum_cb.grid(row=3, column=1, padx=4, sticky="w")

        def update_sum_col(*_):
            headers = self._headers_for_source(self._si["tbl_src"].get())
            opts = [f"{i}: {h}" for i, h in enumerate(headers)]
            self._si_sum_cb["values"] = opts

        self._si["tbl_src"].trace_add("write", update_sum_col)

        tk.Label(panel, text="匹配模式", bg=COLORS["card"], fg=COLORS["text"],
                 font=("Microsoft YaHei UI", 9, "bold"),
                 width=14, anchor="e").grid(row=4, column=0, padx=(12, 4), pady=6)
        mode_frame = tk.Frame(panel, bg=COLORS["card"])
        mode_frame.grid(row=4, column=1, columnspan=3, sticky="w", padx=4)
        for mode_val, mode_lbl in MATCH_MODES:
            ttk.Radiobutton(mode_frame, text=mode_lbl,
                            variable=self._si["match_mode"],
                            value=mode_val).pack(side="left", padx=6)
        panel.columnconfigure(1, weight=1)

    def _build_countif_panel(self):
        panel = make_card(self._param_container)
        self._param_panels["COUNTIF"] = panel

        tk.Label(panel, text="COUNTIF 参数", bg=COLORS["card"],
                 fg=COLORS["text"], font=("Microsoft YaHei UI", 10, "bold")).grid(
                     row=0, column=0, columnspan=4, padx=12, pady=(10, 4), sticky="w")

        self._ci = {}
        self._ci["tbl_src"]      = tk.StringVar()
        self._ci["criteria_col"] = tk.StringVar()
        self._ci["criteria"]     = tk.StringVar()
        self._ci["match_mode"]   = tk.StringVar(value="exact")

        self._make_source_row(panel, 1, "数据表格", self._ci["tbl_src"],
                              self._ci["criteria_col"], "统计列")

        tk.Label(panel, text="条件值 *", bg=COLORS["card"], fg=COLORS["text"],
                 font=("Microsoft YaHei UI", 9, "bold"),
                 width=14, anchor="e").grid(row=2, column=0, padx=(12, 4), pady=6)
        ttk.Entry(panel, textvariable=self._ci["criteria"], width=30).grid(
            row=2, column=1, padx=4, sticky="w")

        tk.Label(panel, text="匹配模式", bg=COLORS["card"], fg=COLORS["text"],
                 font=("Microsoft YaHei UI", 9, "bold"),
                 width=14, anchor="e").grid(row=3, column=0, padx=(12, 4), pady=6)
        mode_frame = tk.Frame(panel, bg=COLORS["card"])
        mode_frame.grid(row=3, column=1, columnspan=3, sticky="w", padx=4)
        for mode_val, mode_lbl in MATCH_MODES:
            ttk.Radiobutton(mode_frame, text=mode_lbl,
                            variable=self._ci["match_mode"],
                            value=mode_val).pack(side="left", padx=6)
        panel.columnconfigure(1, weight=1)

    def _build_sumifs_panel(self):
        panel = make_card(self._param_container)
        self._param_panels["SUMIFS"] = panel

        tk.Label(panel, text="SUMIFS 参数（最多 4 个条件）", bg=COLORS["card"],
                 fg=COLORS["text"], font=("Microsoft YaHei UI", 10, "bold")).grid(
                     row=0, column=0, columnspan=6, padx=12, pady=(10, 4), sticky="w")

        self._sifs = {}
        self._sifs["tbl_src"] = tk.StringVar()
        self._sifs["sum_col"] = tk.StringVar()

        # 数据源
        tk.Label(panel, text="数据表格", bg=COLORS["card"], fg=COLORS["text"],
                 font=("Microsoft YaHei UI", 9, "bold"),
                 width=12, anchor="e").grid(row=1, column=0, padx=(12, 4), pady=6)
        src_cb = ttk.Combobox(panel, textvariable=self._sifs["tbl_src"],
                               width=34, state="readonly")
        src_cb["values"]    = self._source_names()
        src_cb._is_source_cb = True
        src_cb.grid(row=1, column=1, padx=4, columnspan=2, sticky="w")

        tk.Label(panel, text="求和列", bg=COLORS["card"], fg=COLORS["text"],
                 font=("Microsoft YaHei UI", 9, "bold"),
                 width=8, anchor="e").grid(row=1, column=3, padx=4)
        self._sifs_sum_cb = ttk.Combobox(panel, textvariable=self._sifs["sum_col"],
                                          width=18, state="readonly")
        self._sifs_sum_cb.grid(row=1, column=4, padx=4)

        def update_sifs_cols(*_):
            headers = self._headers_for_source(self._sifs["tbl_src"].get())
            opts = [f"{i}: {h}" for i, h in enumerate(headers)]
            self._sifs_sum_cb["values"] = opts
            for i in range(4):
                self._sifs[f"cond{i}_col_cb"]["values"] = opts
        self._sifs["tbl_src"].trace_add("write", update_sifs_cols)

        # 4 个条件行
        cond_hdr = tk.Frame(panel, bg=COLORS["card"])
        cond_hdr.grid(row=2, column=0, columnspan=6, padx=12, pady=(8, 2), sticky="w")
        for txt, w in [("条件列", 20), ("条件值", 20), ("匹配模式", 16)]:
            tk.Label(cond_hdr, text=txt, bg=COLORS["card"], fg=COLORS["text_sub"],
                     font=("Microsoft YaHei UI", 8), width=w).pack(side="left", padx=4)

        for i in range(4):
            col_var  = tk.StringVar()
            crit_var = tk.StringVar()
            mode_var = tk.StringVar(value="exact")
            self._sifs[f"cond{i}_col"]  = col_var
            self._sifs[f"cond{i}_crit"] = crit_var
            self._sifs[f"cond{i}_mode"] = mode_var

            row_f = tk.Frame(panel, bg=COLORS["card"])
            row_f.grid(row=3+i, column=0, columnspan=6, padx=12, pady=2, sticky="w")

            tk.Label(row_f, text=f"条件 {i+1}", bg=COLORS["card"],
                     fg=COLORS["text_sub"], font=("Microsoft YaHei UI", 8),
                     width=6).pack(side="left")
            col_cb = ttk.Combobox(row_f, textvariable=col_var, width=22, state="readonly")
            col_cb.pack(side="left", padx=4)
            self._sifs[f"cond{i}_col_cb"] = col_cb
            ttk.Entry(row_f, textvariable=crit_var, width=20).pack(side="left", padx=4)
            mode_cb = ttk.Combobox(row_f, textvariable=mode_var, width=12,
                                    values=[m[0] for m in MATCH_MODES], state="readonly")
            mode_cb.pack(side="left", padx=4)

        panel.columnconfigure(1, weight=1)

    def _reset_params(self):
        self._rows_cache.clear()

    # ── 执行函数 ───────────────────────────────────────────────────────────────

    def _collect_ui_params(self, fname: str) -> dict:
        """在主线程中收集所有 tkinter 控件的当前值，线程只接收纯 Python 数据。"""
        p = {"fname": fname}
        def _lb_sel(lb_key, var_dict):
            lb  = var_dict[lb_key]
            sel = list(lb.curselection())
            return [int(lb.get(i).split(":")[0]) for i in sel]

        if fname == "VLOOKUP":
            p["lv_src"]     = self._vl["lv_src"].get()
            p["tbl_src"]    = self._vl["tbl_src"].get()
            p["lv_col"]     = self._col_idx_from_var(self._vl["lv_col"])
            p["key_col"]    = self._col_idx_from_var(self._vl["key_col"])
            p["match_mode"] = self._vl["match_mode"].get()
            p["not_found"]  = self._vl["not_found"].get()
            p["ret_cols"]   = _lb_sel("ret_lb", self._vl)
        elif fname == "XLOOKUP":
            p["lv_src"]     = self._xl["lv_src"].get()
            p["search_src"] = self._xl["search_src"].get()
            p["return_src"] = self._xl["return_src"].get()
            p["lv_col"]     = self._col_idx_from_var(self._xl["lv_col"])
            p["search_col"] = self._col_idx_from_var(self._xl["search_col"])
            p["match_mode"] = self._xl["match_mode"].get()
            p["not_found"]  = self._xl["not_found"].get()
            p["ret_cols"]   = _lb_sel("ret_lb", self._xl)
        elif fname == "INDEX/MATCH":
            p["lv_src"]     = self._im["lv_src"].get()
            p["match_src"]  = self._im["match_src"].get()
            p["index_src"]  = self._im["index_src"].get()
            p["lv_col"]     = self._col_idx_from_var(self._im["lv_col"])
            p["match_col"]  = self._col_idx_from_var(self._im["match_col"])
            p["index_col"]  = self._col_idx_from_var(self._im["index_col"])
            p["match_mode"] = self._im["match_mode"].get()
            p["not_found"]  = self._im["not_found"].get()
        elif fname == "SUMIF":
            p["tbl_src"]      = self._si["tbl_src"].get()
            p["criteria_col"] = self._col_idx_from_var(self._si["criteria_col"])
            p["sum_col"]      = self._col_idx_from_var(self._si["sum_col"])
            p["criteria"]     = self._si["criteria"].get().strip()
            p["match_mode"]   = self._si["match_mode"].get()
        elif fname == "COUNTIF":
            p["tbl_src"]      = self._ci["tbl_src"].get()
            p["criteria_col"] = self._col_idx_from_var(self._ci["criteria_col"])
            p["criteria"]     = self._ci["criteria"].get().strip()
            p["match_mode"]   = self._ci["match_mode"].get()
        elif fname == "SUMIFS":
            p["tbl_src"] = self._sifs["tbl_src"].get()
            p["sum_col"] = self._col_idx_from_var(self._sifs["sum_col"])
            conds = []
            for i in range(4):
                crit = self._sifs[f"cond{i}_crit"].get().strip()
                if not crit:
                    continue
                conds.append({
                    "col":      self._col_idx_from_var(self._sifs[f"cond{i}_col"]),
                    "criteria": crit,
                    "mode":     self._sifs[f"cond{i}_mode"].get(),
                })
            p["conds"] = conds
        return p

    def _run_function(self):
        fname = self._func_var.get()
        self._run_info.config(text="执行中...", fg=COLORS["warning"])
        self.update_idletasks()

        # 在主线程收集所有 UI 值，子线程不接触任何 tkinter 控件
        try:
            params = self._collect_ui_params(fname)
        except Exception as ex:
            self._run_info.config(text=f"错误: {ex}", fg=COLORS["error"])
            messagebox.showerror("参数错误", str(ex), parent=self)
            return

        def worker():
            try:
                headers, rows, status_col = self._execute(params)
                self.after(0, lambda: self._run_info.config(
                    text=f"完成  共 {len(rows)} 行结果", fg=COLORS["success"]))
                if self.on_results:
                    self.after(0, lambda: self.on_results(headers, rows, fname))
            except Exception as ex:
                self.after(0, lambda e=ex: [
                    self._run_info.config(text=f"错误: {e}", fg=COLORS["error"]),
                    messagebox.showerror("执行失败", str(e), parent=self),
                ])

        threading.Thread(target=worker, daemon=True).start()

    def _execute(self, p: dict):
        """线程安全：接收主线程已收集好的纯 Python 参数字典，不访问任何 tkinter 控件。"""
        fname  = p["fname"]
        engine = FormulaEngine()

        def get_rows(choice: str) -> list[list[str]]:
            src = self._source_by_choice(choice)
            if not src:
                raise ValueError(f"请选择数据源（当前: {choice!r}）")
            if not src.get("synced_at"):
                raise ValueError(f"数据源「{src['name']}」尚未同步，请先同步")
            return self._get_rows(src["id"])

        if fname == "VLOOKUP":
            lv_rows  = get_rows(p["lv_src"])
            tbl_rows = get_rows(p["tbl_src"])
            lv_col, key_col = p["lv_col"], p["key_col"]
            mode, not_f, ret_cols = p["match_mode"], p["not_found"], p["ret_cols"]
            if not ret_cols:
                raise ValueError("请至少选择一个返回列")
            lv_list  = [str(r[lv_col]) if lv_col < len(r) else "" for r in lv_rows[1:]]
            results  = engine.vlookup(lv_list, tbl_rows, key_col, ret_cols, mode, not_f)
            tbl_hdr  = tbl_rows[0] if tbl_rows else []
            ret_hdrs = [tbl_hdr[c] if c < len(tbl_hdr) else f"列{c}" for c in ret_cols]
            lv_hdr   = lv_rows[0][lv_col] if lv_rows and lv_col < len(lv_rows[0]) else "查找值"
            headers  = [lv_hdr] + ret_hdrs + ["匹配状态"]
            out_rows = [[r["__lookup_value__"]] +
                        [r.get(f"col_{c}", not_f) for c in ret_cols] +
                        [r["__status__"]] for r in results]
            return headers, out_rows, len(headers) - 1

        elif fname == "XLOOKUP":
            lv_rows     = get_rows(p["lv_src"])
            search_rows = get_rows(p["search_src"])
            return_rows = get_rows(p["return_src"])
            lv_col, search_col = p["lv_col"], p["search_col"]
            mode, not_f, ret_cols = p["match_mode"], p["not_found"], p["ret_cols"]
            if not ret_cols:
                raise ValueError("请至少选择一个返回列")
            lv_list  = [str(r[lv_col]) if lv_col < len(r) else "" for r in lv_rows[1:]]
            results  = engine.xlookup(lv_list, search_rows, search_col,
                                      return_rows, ret_cols, mode, not_f)
            ret_hdr  = return_rows[0] if return_rows else []
            ret_hdrs = [ret_hdr[c] if c < len(ret_hdr) else f"列{c}" for c in ret_cols]
            lv_hdr   = lv_rows[0][lv_col] if lv_rows and lv_col < len(lv_rows[0]) else "查找值"
            headers  = [lv_hdr] + ret_hdrs + ["匹配状态"]
            out_rows = [[r["__lookup_value__"]] +
                        [r.get(f"col_{c}", not_f) for c in ret_cols] +
                        [r["__status__"]] for r in results]
            return headers, out_rows, len(headers) - 1

        elif fname == "INDEX/MATCH":
            lv_rows    = get_rows(p["lv_src"])
            match_rows = get_rows(p["match_src"])
            index_rows = get_rows(p["index_src"])
            lv_col, match_col, index_col = p["lv_col"], p["match_col"], p["index_col"]
            mode, not_f = p["match_mode"], p["not_found"]
            lv_list  = [str(r[lv_col]) if lv_col < len(r) else "" for r in lv_rows[1:]]
            results  = engine.index_match(lv_list, match_rows, match_col,
                                          index_rows, index_col, mode, not_f)
            idx_hdr  = index_rows[0] if index_rows else []
            ret_name = idx_hdr[index_col] if index_col < len(idx_hdr) else f"列{index_col}"
            lv_hdr   = lv_rows[0][lv_col] if lv_rows and lv_col < len(lv_rows[0]) else "查找值"
            headers  = [lv_hdr, ret_name, "匹配行号", "匹配状态"]
            out_rows = [[r["__lookup_value__"], r["result"],
                         str(r.get("_row_idx", "")), r["__status__"]] for r in results]
            return headers, out_rows, 3

        elif fname == "SUMIF":
            tbl_rows     = get_rows(p["tbl_src"])
            criteria_col = p["criteria_col"]
            sum_col      = p["sum_col"]
            criteria     = p["criteria"]
            mode         = p["match_mode"]
            if not criteria:
                raise ValueError("请填写条件值")
            result    = engine.sumif(tbl_rows, criteria_col, criteria, sum_col, mode)
            tbl_hdr   = tbl_rows[0] if tbl_rows else []
            sum_name  = tbl_hdr[sum_col] if sum_col < len(tbl_hdr) else f"列{sum_col}"
            crit_name = tbl_hdr[criteria_col] if criteria_col < len(tbl_hdr) else f"列{criteria_col}"
            headers   = [crit_name, sum_name] + [h for h in tbl_hdr if h != sum_name and h != crit_name]
            sum_row   = [f"[合计]  条件: {criteria}", str(result["sum"])] + [""]*(len(headers)-2)
            cnt_row   = [f"[匹配行数: {result['count']}]", ""] + [""]*(len(headers)-2)
            out_rows  = [sum_row, cnt_row]
            for r in result["matched_rows"]:
                padded = r + [""] * (len(headers) - len(r))
                out_rows.append([padded[criteria_col], padded[sum_col]] +
                                 [v for ci, v in enumerate(padded) if ci != criteria_col and ci != sum_col])
            return headers, out_rows, None

        elif fname == "COUNTIF":
            tbl_rows     = get_rows(p["tbl_src"])
            criteria_col = p["criteria_col"]
            criteria     = p["criteria"]
            mode         = p["match_mode"]
            if not criteria:
                raise ValueError("请填写条件值")
            result    = engine.countif(tbl_rows, criteria_col, criteria, mode)
            tbl_hdr   = tbl_rows[0] if tbl_rows else []
            crit_name = tbl_hdr[criteria_col] if criteria_col < len(tbl_hdr) else f"列{criteria_col}"
            headers   = tbl_hdr if tbl_hdr else [f"列{i}" for i in range(
                max(len(r) for r in result["matched_rows"]) if result["matched_rows"] else 1)]
            sum_row   = [f"[COUNTIF 结果: {result['count']}  条件: {crit_name}={criteria}]"]
            sum_row  += [""] * (len(headers) - 1)
            out_rows  = [sum_row] + [r + [""]*(len(headers)-len(r)) for r in result["matched_rows"]]
            return headers, out_rows, None

        elif fname == "SUMIFS":
            tbl_rows = get_rows(p["tbl_src"])
            sum_col  = p["sum_col"]
            conds    = p["conds"]
            if not conds:
                raise ValueError("请至少填写一个条件")
            result   = engine.sumifs(tbl_rows, sum_col, conds)
            tbl_hdr  = tbl_rows[0] if tbl_rows else []
            headers  = tbl_hdr if tbl_hdr else [f"列{i}" for i in range(
                max(len(r) for r in result["matched_rows"]) if result["matched_rows"] else 1)]
            cond_str = "  &  ".join(f"条件{i+1}={c['criteria']}" for i, c in enumerate(conds))
            sum_row  = [f"[SUMIFS 合计: {result['sum']}  匹配行: {result['count']}  {cond_str}]"]
            sum_row += [""] * (len(headers) - 1)
            out_rows = [sum_row] + [r + [""]*(len(headers)-len(r)) for r in result["matched_rows"]]
            return headers, out_rows, None

        raise ValueError(f"未知函数类型: {fname}")


# ═════════════════════════════════════════════════════════════════════════════
# GUI 面板 3 — 结果预览
# ═════════════════════════════════════════════════════════════════════════════

class ResultsPanel(tk.Frame):
    """显示函数执行结果，支持搜索过滤、导出和写回飞书"""

    def __init__(self, master, cache: DataCache = None,
                 api: FeishuAPIClient = None, **kwargs):
        super().__init__(master, bg=COLORS["bg"], **kwargs)
        self.cache = cache
        self.api   = api
        self._headers: list[str] = []
        self._rows: list[list[str]] = []
        self._filtered: list[list[str]] = []
        self._func_type: str = ""
        self._status_col: int | None = None
        self._build_ui()

    def _build_ui(self):
        # ── 顶栏 ──────────────────────────────────────────────────────────────
        top = tk.Frame(self, bg=COLORS["bg"])
        top.pack(fill="x", padx=16, pady=(12, 4))

        tk.Label(top, text="执行结果", bg=COLORS["bg"],
                 fg=COLORS["text"], font=("Microsoft YaHei UI", 13, "bold")).pack(side="left")

        self._result_info = tk.Label(top, text="尚未执行任何函数", bg=COLORS["bg"],
                                     fg=COLORS["text_sub"], font=("Microsoft YaHei UI", 9))
        self._result_info.pack(side="left", padx=16)

        # 搜索栏
        search_frame = tk.Frame(top, bg=COLORS["bg"])
        search_frame.pack(side="right")
        tk.Label(search_frame, text="搜索:", bg=COLORS["bg"],
                 fg=COLORS["text_sub"], font=("Microsoft YaHei UI", 9)).pack(side="left")
        self._search_var = tk.StringVar()
        self._search_var.trace_add("write", self._apply_filter)
        ttk.Entry(search_frame, textvariable=self._search_var, width=24).pack(
            side="left", padx=4)

        # ── 状态统计栏 ────────────────────────────────────────────────────────
        stat_bar = tk.Frame(self, bg=COLORS["bg"])
        stat_bar.pack(fill="x", padx=16, pady=2)

        self._stat_match   = tk.Label(stat_bar, text="", bg=COLORS["bg"],
                                       fg=COLORS["success"], font=("Microsoft YaHei UI", 9))
        self._stat_match.pack(side="left", padx=4)
        self._stat_nomatch = tk.Label(stat_bar, text="", bg=COLORS["bg"],
                                       fg=COLORS["error"], font=("Microsoft YaHei UI", 9))
        self._stat_nomatch.pack(side="left", padx=4)

        # ── 结果表格 ──────────────────────────────────────────────────────────
        tbl_card = make_card(self)
        tbl_card.pack(fill="both", expand=True, padx=16, pady=4)

        self.tree = ttk.Treeview(tbl_card, show="headings",
                                  selectmode="extended", height=20)
        vsb  = ttk.Scrollbar(tbl_card, orient="vertical",   command=self.tree.yview)
        hsb  = ttk.Scrollbar(tbl_card, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tbl_card.rowconfigure(0, weight=1)
        tbl_card.columnconfigure(0, weight=1)

        self.tree.tag_configure("match",    background="#E6F7F0")
        self.tree.tag_configure("nomatch",  background="#FFF1F0")
        self.tree.tag_configure("even",     background=COLORS["row_even"])
        self.tree.tag_configure("odd",      background=COLORS["row_odd"])
        self.tree.tag_configure("summary",  background=COLORS["highlight"],
                                font=("Microsoft YaHei UI", 9, "bold"))

        # ── 底部工具栏 ────────────────────────────────────────────────────────
        bot = tk.Frame(self, bg=COLORS["bg"])
        bot.pack(fill="x", padx=16, pady=8)

        make_label_btn(bot, "↓ 导出 Excel", self._export_excel,
                       bg=COLORS["success"]).pack(side="left", padx=4)
        make_label_btn(bot, "↑ 写回飞书", self._open_write_back_dialog,
                       bg=COLORS["primary"]).pack(side="left", padx=4)
        make_label_btn(bot, "⎘ 复制表头+数据", self._copy_tsv,
                       bg=COLORS["warning"]).pack(side="left", padx=4)

        # 显示模式
        self._filter_mode = tk.StringVar(value="all")
        filter_frame = tk.Frame(bot, bg=COLORS["bg"])
        filter_frame.pack(side="right")
        for text, val in [("全部", "all"), ("仅匹配", "match"), ("仅未匹配", "nomatch")]:
            ttk.Radiobutton(filter_frame, text=text, variable=self._filter_mode,
                            value=val,
                            command=self._apply_filter).pack(side="left", padx=4)

    def show_results(self, headers: list[str], rows: list[list[str]],
                     func_type: str, status_col: int | None = None):
        self._headers    = headers
        self._rows       = rows
        self._func_type  = func_type
        self._status_col = status_col
        self._apply_filter()

        total   = len(rows)
        matched = sum(1 for r in rows if status_col is not None
                      and status_col < len(r) and r[status_col] == "MATCH")
        self._result_info.config(
            text=f"函数: {func_type}  |  共 {total} 行"
        )
        if status_col is not None:
            self._stat_match.config(  text=f"匹配: {matched}")
            self._stat_nomatch.config(text=f"未匹配: {total - matched}")
        else:
            self._stat_match.config(  text="")
            self._stat_nomatch.config(text="")

    def _apply_filter(self, *_):
        kw   = self._search_var.get().strip().lower()
        mode = self._filter_mode.get()
        rows = self._rows

        if kw:
            rows = [r for r in rows if any(kw in str(v).lower() for v in r)]

        if mode != "all" and self._status_col is not None:
            sc = self._status_col
            if mode == "match":
                rows = [r for r in rows if sc < len(r) and r[sc] == "MATCH"]
            elif mode == "nomatch":
                rows = [r for r in rows if sc < len(r) and r[sc] != "MATCH"]

        self._filtered = rows
        self._render_table()

    def _render_table(self):
        self.tree.delete(*self.tree.get_children())
        if not self._headers:
            return

        col_ids = [f"c{i}" for i in range(len(self._headers))]
        self.tree["columns"] = col_ids
        for ci, (cid, h) in enumerate(zip(col_ids, self._headers)):
            self.tree.heading(cid, text=h, anchor="w")
            self.tree.column(cid, width=120, minwidth=60, anchor="w")

        for ri, row in enumerate(self._filtered):
            padded = list(row) + [""] * (len(self._headers) - len(row))

            if self._status_col is not None and self._status_col < len(padded):
                status = padded[self._status_col]
                tag    = "match" if status == "MATCH" else "nomatch"
            elif ri == 0 and padded and str(padded[0]).startswith("["):
                tag = "summary"
            else:
                tag = "even" if ri % 2 == 0 else "odd"

            self.tree.insert("", "end", values=padded, tags=(tag,))

    def _export_excel(self):
        if not self._rows:
            messagebox.showinfo("提示", "暂无结果可导出", parent=self)
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx")],
            initialfile=f"函数结果_{self._func_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            parent=self,
        )
        if not path:
            return
        try:
            ExcelExporter.export_results(path, self._headers, self._rows, self._status_col)
            messagebox.showinfo("导出成功", f"已保存至:\n{path}", parent=self)
        except Exception as ex:
            messagebox.showerror("导出失败", str(ex), parent=self)

    def _copy_tsv(self):
        if not self._rows:
            return
        lines = ["\t".join(self._headers)]
        lines += ["\t".join(str(v) for v in r) for r in self._filtered]
        self.clipboard_clear()
        self.clipboard_append("\n".join(lines))
        messagebox.showinfo("已复制", f"已复制 {len(self._filtered)} 行到剪贴板", parent=self)

    def _open_write_back_dialog(self):
        """打开「写回飞书」对话框：选择目标表格和起始位置，写入函数结果"""
        if not self._rows:
            messagebox.showinfo("提示", "暂无结果，请先执行函数", parent=self)
            return
        if not self.cache or not self.api:
            messagebox.showwarning("提示", "未配置 cache/api，无法写回", parent=self)
            return

        dlg = tk.Toplevel(self)
        dlg.title("写回飞书")
        dlg.geometry("620x620")
        dlg.resizable(False, False)
        dlg.grab_set()
        dlg.configure(bg=COLORS["bg"])

        tk.Label(dlg, text="将函数结果写回飞书表格", bg=COLORS["bg"],
                 fg=COLORS["text"],
                 font=("Microsoft YaHei UI", 12, "bold")).pack(pady=(16, 4))
        tk.Label(dlg, text="选择目标数据源（已同步的飞书 Sheet），配置写入起始位置",
                 bg=COLORS["bg"], fg=COLORS["text_sub"],
                 font=("Microsoft YaHei UI", 9)).pack(pady=(0, 8))

        cfg_card = make_card(dlg)
        cfg_card.pack(fill="x", padx=20, pady=4)

        # 目标数据源
        tk.Label(cfg_card, text="目标数据源", bg=COLORS["card"], fg=COLORS["text"],
                 font=("Microsoft YaHei UI", 9, "bold"),
                 width=12, anchor="e").grid(row=0, column=0, padx=(12, 4), pady=10)
        sources = self.cache.list_sources()
        src_names = [f"{s['id']}: {s['name']} [{s['sheet_title'] or s['sheet_id']}]"
                     for s in sources]
        tgt_src_var = tk.StringVar()
        src_cb = ttk.Combobox(cfg_card, textvariable=tgt_src_var,
                               values=src_names, width=40, state="readonly")
        src_cb.grid(row=0, column=1, columnspan=2, padx=4, pady=10, sticky="w")

        # 起始行
        tk.Label(cfg_card, text="起始行(0=表头)", bg=COLORS["card"], fg=COLORS["text"],
                 font=("Microsoft YaHei UI", 9, "bold"),
                 width=12, anchor="e").grid(row=1, column=0, padx=(12, 4), pady=8)
        start_row_var = tk.StringVar(value="1")
        ttk.Entry(cfg_card, textvariable=start_row_var, width=8).grid(
            row=1, column=1, padx=4, sticky="w")
        tk.Label(cfg_card, text="（1=跳过表头，从第二行开始写）", bg=COLORS["card"],
                 fg=COLORS["text_sub"], font=("Microsoft YaHei UI", 8)).grid(
                     row=1, column=2, padx=4, sticky="w")

        # 起始列
        tk.Label(cfg_card, text="起始列(0=A列)", bg=COLORS["card"], fg=COLORS["text"],
                 font=("Microsoft YaHei UI", 9, "bold"),
                 width=12, anchor="e").grid(row=2, column=0, padx=(12, 4), pady=8)
        start_col_var = tk.StringVar(value="0")
        ttk.Entry(cfg_card, textvariable=start_col_var, width=8).grid(
            row=2, column=1, padx=4, sticky="w")
        tk.Label(cfg_card, text="（若只需写一列，可配合「写入列选择」使用）",
                 bg=COLORS["card"], fg=COLORS["text_sub"],
                 font=("Microsoft YaHei UI", 8)).grid(row=2, column=2, padx=4, sticky="w")

        # 选择写入哪些结果列
        tk.Label(cfg_card, text="写入列（多选）", bg=COLORS["card"], fg=COLORS["text"],
                 font=("Microsoft YaHei UI", 9, "bold"),
                 width=12, anchor="e").grid(row=3, column=0, padx=(12, 4), pady=8)
        col_frame = tk.Frame(cfg_card, bg=COLORS["card"])
        col_frame.grid(row=3, column=1, columnspan=2, padx=4, pady=8, sticky="w")
        col_lb = tk.Listbox(col_frame, selectmode="multiple", height=5, width=42)
        col_sb = ttk.Scrollbar(col_frame, command=col_lb.yview)
        col_lb.configure(yscrollcommand=col_sb.set)
        col_lb.pack(side="left")
        col_sb.pack(side="left", fill="y")

        # 排除状态列，其余均可选
        writable_cols: list[tuple[int, str]] = []
        for ci, h in enumerate(self._headers):
            if h not in ("匹配状态", "__status__"):
                writable_cols.append((ci, h))
                col_lb.insert("end", f"{ci}: {h}")
        # 默认全选
        col_lb.select_set(0, "end")

        # 是否跳过未匹配行
        skip_nomatch_var = tk.BooleanVar(value=True)
        tk.Checkbutton(cfg_card, text="跳过未匹配行（状态=NOT_FOUND）",
                       variable=skip_nomatch_var,
                       bg=COLORS["card"], font=("Microsoft YaHei UI", 9)).grid(
                           row=4, column=0, columnspan=3, padx=12, pady=6, sticky="w")

        cfg_card.columnconfigure(2, weight=1)

        # 预览
        preview_lbl = tk.Label(dlg, text="写入预览（前 5 行）", bg=COLORS["bg"],
                               fg=COLORS["text_sub"],
                               font=("Microsoft YaHei UI", 8)).pack(anchor="w", padx=20) if False else None
        tk.Label(dlg, text="写入预览（前 5 行）", bg=COLORS["bg"],
                 fg=COLORS["text_sub"],
                 font=("Microsoft YaHei UI", 8)).pack(anchor="w", padx=20, pady=(8, 2))

        prev_frame = make_card(dlg)
        prev_frame.pack(fill="x", padx=20, pady=2)
        prev_tv = ttk.Treeview(prev_frame, show="headings", height=5)
        prev_tv.pack(fill="x", padx=4, pady=4)

        def update_preview(*_):
            sel_idxs = list(col_lb.curselection())
            if not sel_idxs:
                return
            chosen_cols = [writable_cols[i] for i in sel_idxs]
            prev_tv["columns"] = [f"c{i}" for i in range(len(chosen_cols))]
            for ci, (_, h) in enumerate(chosen_cols):
                prev_tv.heading(f"c{ci}", text=h)
                prev_tv.column(f"c{ci}", width=100)
            prev_tv.delete(*prev_tv.get_children())
            shown = 0
            for row in self._rows:
                if self._status_col is not None and self._status_col < len(row):
                    if skip_nomatch_var.get() and row[self._status_col] == "NOT_FOUND":
                        continue
                preview_row = [row[c] if c < len(row) else "" for c, _ in chosen_cols]
                prev_tv.insert("", "end", values=preview_row)
                shown += 1
                if shown >= 5:
                    break

        col_lb.bind("<<ListboxSelect>>", update_preview)
        skip_nomatch_var.trace_add("write", update_preview)
        update_preview()

        # 状态
        wb_status = tk.StringVar(value="")
        tk.Label(dlg, textvariable=wb_status, bg=COLORS["bg"],
                 fg=COLORS["text_sub"], font=("Microsoft YaHei UI", 8),
                 wraplength=560).pack(pady=4)

        # 按钮
        btn_row = tk.Frame(dlg, bg=COLORS["bg"])
        btn_row.pack(pady=8)

        def do_write():
            tgt = tgt_src_var.get()
            if not tgt:
                messagebox.showwarning("提示", "请选择目标数据源", parent=dlg)
                return
            try:
                start_row = int(start_row_var.get().strip())
                start_col = int(start_col_var.get().strip())
            except ValueError:
                messagebox.showwarning("提示", "起始行/列请输入整数", parent=dlg)
                return

            sel_idxs = list(col_lb.curselection())
            if not sel_idxs:
                messagebox.showwarning("提示", "请至少选择一个写入列", parent=dlg)
                return
            chosen_cols = [writable_cols[i][0] for i in sel_idxs]

            # 过滤行
            write_rows = []
            for row in self._rows:
                if self._status_col is not None and self._status_col < len(row):
                    if skip_nomatch_var.get() and row[self._status_col] == "NOT_FOUND":
                        continue
                write_rows.append([row[c] if c < len(row) else "" for c in chosen_cols])

            if not write_rows:
                messagebox.showinfo("提示", "过滤后没有可写入的行", parent=dlg)
                return

            # 解析目标源
            try:
                sid  = int(tgt.split(":")[0])
                src  = self.cache.get_source(sid)
            except Exception:
                messagebox.showerror("错误", "无法解析目标数据源", parent=dlg)
                return
            if not src:
                messagebox.showerror("错误", "目标数据源已不存在，请重新选择", parent=dlg)
                return

            wb_status.set(f"正在写入 {len(write_rows)} 行 × {len(chosen_cols)} 列...")
            dlg.update_idletasks()

            n = len(write_rows)
            src_name = src["name"]

            def worker():
                try:
                    self.api.write_values(
                        src["token"], src["sheet_id"], write_rows,
                        start_row=start_row, start_col=start_col,
                        use_open_api=bool(src["use_open_api"]),
                    )
                    def _ok():
                        if not dlg.winfo_exists(): return
                        wb_status.set(f"写入成功！共 {n} 行")
                        messagebox.showinfo("成功", f"已写入 {n} 行到\n{src_name}", parent=dlg)
                    dlg.after(0, _ok)
                except Exception as ex:
                    def _err(e=ex):
                        if not dlg.winfo_exists(): return
                        wb_status.set(f"写入失败: {e}")
                        messagebox.showerror("写入失败", str(e), parent=dlg)
                    dlg.after(0, _err)

            threading.Thread(target=worker, daemon=True).start()

        make_label_btn(btn_row, "↑ 执行写回", do_write,
                       padx=16, pady=6, font_size=10).pack(side="left", padx=6)
        make_label_btn(btn_row, "取消", dlg.destroy,
                       bg=COLORS["text_sub"]).pack(side="left", padx=6)


# ═════════════════════════════════════════════════════════════════════════════
# GUI 面板 4 — 在线编辑器
# ═════════════════════════════════════════════════════════════════════════════

class OnlineEditorPanel(tk.Frame):
    """
    在线表格编辑器
    - 从缓存（或直接拉取）加载飞书表格数据到可编辑网格
    - 双击单元格进行编辑，修改后高亮为蓝色
    - "保存到飞书"按钮批量写回所有脏单元格
    - 支持在指定列粘贴一列数据（常用于把 VLOOKUP 结果填到目标列）
    """

    def __init__(self, master, cache: DataCache, api: FeishuAPIClient,
                 **kwargs):
        super().__init__(master, bg=COLORS["bg"], **kwargs)
        self.cache = cache
        self.api   = api

        self._source: dict | None = None   # 当前加载的数据源
        self._data:   list[list[str]] = [] # 包含表头的全量数据（本地副本）
        self._dirty:  dict[tuple[int,int], str] = {}  # {(row_idx, col_idx): new_val}
        self._col_count = 0
        self._editing_entry: tk.Entry | None = None
        self._iid_to_row: dict[str, int] = {}

        self._build_ui()
        self._reload_source_list()

    # ── UI 构建 ───────────────────────────────────────────────────────────────

    def _build_ui(self):
        # 顶栏：数据源选择 + 操作按钮
        top = tk.Frame(self, bg=COLORS["bg"])
        top.pack(fill="x", padx=16, pady=(12, 4))

        tk.Label(top, text="在线编辑", bg=COLORS["bg"],
                 fg=COLORS["text"],
                 font=("Microsoft YaHei UI", 13, "bold")).pack(side="left")

        ctrl = tk.Frame(top, bg=COLORS["bg"])
        ctrl.pack(side="right")

        tk.Label(ctrl, text="数据源:", bg=COLORS["bg"],
                 fg=COLORS["text_sub"],
                 font=("Microsoft YaHei UI", 9)).pack(side="left")
        self._src_var = tk.StringVar()
        self._src_cb = ttk.Combobox(ctrl, textvariable=self._src_var,
                                     width=32, state="readonly")
        self._src_cb.pack(side="left", padx=4)
        self._src_cb.bind("<<ComboboxSelected>>", self._load_source)

        make_label_btn(ctrl, "↻ 重新拉取", self._pull_from_feishu,
                       bg=COLORS["success"]).pack(side="left", padx=4)

        # 状态行
        status_bar = tk.Frame(self, bg=COLORS["bg"])
        status_bar.pack(fill="x", padx=16, pady=2)
        self._status_var = tk.StringVar(value="请选择数据源")
        tk.Label(status_bar, textvariable=self._status_var,
                 bg=COLORS["bg"], fg=COLORS["text_sub"],
                 font=("Microsoft YaHei UI", 8)).pack(side="left")
        self._dirty_var = tk.StringVar(value="")
        self._dirty_lbl = tk.Label(status_bar, textvariable=self._dirty_var,
                                    bg=COLORS["bg"], fg=COLORS["warning"],
                                    font=("Microsoft YaHei UI", 8, "bold"))
        self._dirty_lbl.pack(side="left", padx=12)

        # 编辑提示
        tk.Label(status_bar,
                 text="双击单元格编辑  |  Enter/Tab 确认  |  Esc 取消",
                 bg=COLORS["bg"], fg=COLORS["text_sub"],
                 font=("Microsoft YaHei UI", 8)).pack(side="right")

        # 表格
        tbl_card = make_card(self)
        tbl_card.pack(fill="both", expand=True, padx=16, pady=4)

        self.tree = ttk.Treeview(tbl_card, show="headings",
                                  selectmode="browse", height=24)
        vsb  = ttk.Scrollbar(tbl_card, orient="vertical",   command=self.tree.yview)
        hsb  = ttk.Scrollbar(tbl_card, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tbl_card.rowconfigure(0, weight=1)
        tbl_card.columnconfigure(0, weight=1)

        self.tree.tag_configure("header_row", background=COLORS["primary"],
                                font=("Microsoft YaHei UI", 8, "bold"))
        self.tree.tag_configure("even",   background=COLORS["row_even"])
        self.tree.tag_configure("odd",    background=COLORS["row_odd"])
        self.tree.tag_configure("dirty",  background="#FFF3CD",
                                foreground="#856404")

        self.tree.bind("<Double-1>", self._on_dblclick)
        self.tree.bind("<Button-1>", self._commit_editing)

        # 底部操作栏
        bot = tk.Frame(self, bg=COLORS["bg"])
        bot.pack(fill="x", padx=16, pady=8)

        make_label_btn(bot, "↑ 保存修改到飞书", self._save_dirty,
                       bg=COLORS["success"], padx=16, pady=6,
                       font_size=10).pack(side="left", padx=4)
        make_label_btn(bot, "✗ 撤销所有修改", self._discard_dirty,
                       bg=COLORS["error"]).pack(side="left", padx=4)

        # 粘贴列工具
        paste_sep = tk.Frame(bot, bg=COLORS["border"], width=1)
        paste_sep.pack(side="left", fill="y", padx=8)

        tk.Label(bot, text="粘贴一列到列号:",
                 bg=COLORS["bg"], fg=COLORS["text_sub"],
                 font=("Microsoft YaHei UI", 9)).pack(side="left")
        self._paste_col_var = tk.StringVar(value="0")
        ttk.Entry(bot, textvariable=self._paste_col_var, width=5).pack(
            side="left", padx=4)
        tk.Label(bot, text="起始行(含表头=0):",
                 bg=COLORS["bg"], fg=COLORS["text_sub"],
                 font=("Microsoft YaHei UI", 9)).pack(side="left", padx=(8, 0))
        self._paste_row_var = tk.StringVar(value="1")
        ttk.Entry(bot, textvariable=self._paste_row_var, width=5).pack(
            side="left", padx=4)
        make_label_btn(bot, "粘贴剪贴板列", self._paste_col_from_clipboard,
                       bg=COLORS["warning"]).pack(side="left", padx=4)

    # ── 数据源 ────────────────────────────────────────────────────────────────

    def _reload_source_list(self):
        sources = self.cache.list_sources()
        names   = [f"{s['id']}: {s['name']} [{s['sheet_title'] or s['sheet_id']}]"
                   for s in sources]
        self._src_cb["values"] = names

    def _source_id_from_choice(self) -> int | None:
        v = self._src_var.get()
        try:
            return int(v.split(":")[0])
        except (ValueError, IndexError):
            return None

    def _load_source(self, event=None):
        sid = self._source_id_from_choice()
        if sid is None:
            return
        src = self.cache.get_source(sid)
        if not src:
            return
        if not src.get("synced_at"):
            if messagebox.askyesno("未同步", "该数据源尚未同步，立即从飞书拉取？",
                                   parent=self):
                self._pull_from_feishu()
            return
        self._source = src
        rows = self.cache.get_all_rows(sid)
        self._data   = [list(r) for r in rows]
        self._dirty  = {}
        self._render_grid()
        self._status_var.set(
            f"已加载: {src['name']}  |  {len(rows)} 行 × {src['col_count']} 列  "
            f"|  上次同步: {src['synced_at']}")
        self._dirty_var.set("")

    def _pull_from_feishu(self):
        sid = self._source_id_from_choice()
        if sid is None:
            messagebox.showinfo("提示", "请先选择数据源", parent=self)
            return
        src = self.cache.get_source(sid)
        if not src:
            return
        self._status_var.set("正在从飞书拉取数据...")
        self.update_idletasks()

        def worker():
            try:
                rows = self.api.fetch_values(
                    src["token"], src["sheet_id"],
                    use_open_api=bool(src["use_open_api"])
                )
                self.cache.store_rows(sid, rows)
                self.after(0, self._load_source)
                self.after(0, lambda: self._status_var.set(
                    f"拉取完成: {len(rows)} 行"))
            except Exception as ex:
                self.after(0, lambda e=ex: [
                    self._status_var.set(f"拉取失败: {e}"),
                    messagebox.showerror("拉取失败", str(e), parent=self),
                ])

        threading.Thread(target=worker, daemon=True).start()

    # ── 渲染网格 ──────────────────────────────────────────────────────────────

    def _render_grid(self):
        self.tree.delete(*self.tree.get_children())
        if not self._data:
            return
        headers = self._data[0]
        self._col_count = len(headers)

        col_ids = [f"c{i}" for i in range(len(headers))]
        self.tree["columns"] = col_ids
        for ci, (cid, h) in enumerate(zip(col_ids, headers)):
            self.tree.heading(cid, text=f"{ci} | {h}", anchor="w")
            self.tree.column(cid, width=120, minwidth=60, anchor="w")

        # 存储 item_id -> row_idx 映射（用于编辑定位）
        self._iid_to_row: dict[str, int] = {}

        for ri, row in enumerate(self._data[1:], start=1):
            padded = list(row) + [""] * (len(headers) - len(row))
            # 检查该行是否有脏单元格
            row_dirty = any((ri, ci) in self._dirty for ci in range(len(headers)))
            tag = "dirty" if row_dirty else ("even" if ri % 2 == 0 else "odd")
            iid = self.tree.insert("", "end", values=padded, tags=(tag,))
            self._iid_to_row[iid] = ri

    def _refresh_row_tag(self, iid: str):
        ri = self._iid_to_row.get(iid)
        if ri is None:
            return
        row_dirty = any((ri, ci) in self._dirty for ci in range(self._col_count))
        tag = "dirty" if row_dirty else ("even" if ri % 2 == 0 else "odd")
        self.tree.item(iid, tags=(tag,))

    # ── 单元格双击编辑 ────────────────────────────────────────────────────────

    def _commit_editing(self, event=None):
        """点击其他区域时提交当前正在编辑的 Entry"""
        if self._editing_entry and self._editing_entry.winfo_exists():
            self._editing_entry.event_generate("<Return>")

    def _on_dblclick(self, event):
        # 先提交任何正在进行的编辑
        self._commit_editing()

        item = self.tree.identify_row(event.y)
        col  = self.tree.identify_column(event.x)
        if not item or not col:
            return
        bbox = self.tree.bbox(item, col)
        if not bbox:
            return

        x, y, w, h = bbox
        col_idx = int(col[1:]) - 1   # "#1" → 0
        row_idx = self._iid_to_row.get(item)
        if row_idx is None:
            return

        # 确保本地数据足够长
        while len(self._data) <= row_idx:
            self._data.append([""] * self._col_count)
        while len(self._data[row_idx]) <= col_idx:
            self._data[row_idx].append("")

        cur_val = self._data[row_idx][col_idx]

        var   = tk.StringVar(value=cur_val)
        entry = tk.Entry(self.tree, textvariable=var,
                         bg="white", fg=COLORS["primary"],
                         insertbackground=COLORS["primary"],
                         relief="solid", bd=1,
                         font=("Microsoft YaHei UI", 9))
        entry.place(x=x, y=y, width=max(w, 80), height=h)
        entry.focus_set()
        entry.select_range(0, "end")
        self._editing_entry = entry

        def commit(e=None):
            if not entry.winfo_exists():
                return
            new_val = var.get()
            old_val = self._data[row_idx][col_idx]

            if new_val != old_val:
                self._data[row_idx][col_idx] = new_val
                self._dirty[(row_idx, col_idx)] = new_val
                # 更新 Treeview 显示
                vals = list(self.tree.item(item, "values"))
                while len(vals) <= col_idx:
                    vals.append("")
                vals[col_idx] = new_val
                self.tree.item(item, values=vals)
                self._refresh_row_tag(item)
                self._dirty_var.set(f"● 未保存修改: {len(self._dirty)} 个单元格")
            entry.destroy()
            self._editing_entry = None

        def cancel(e=None):
            if entry.winfo_exists():
                entry.destroy()
            self._editing_entry = None

        entry.bind("<Return>", commit)
        entry.bind("<Tab>",    commit)
        entry.bind("<Escape>", cancel)
        # 注意：不绑定 FocusOut，否则点别处时会与 _commit_editing 冲突

    # ── 保存 / 撤销 ──────────────────────────────────────────────────────────

    def _save_dirty(self):
        if not self._dirty:
            messagebox.showinfo("提示", "没有需要保存的修改", parent=self)
            return
        if not self._source:
            messagebox.showwarning("提示", "未加载数据源", parent=self)
            return

        # 将所有脏单元格按行分组，构建最小矩形区域逐行写入
        rows_to_write: dict[int, dict[int, str]] = {}
        for (ri, ci), val in self._dirty.items():
            rows_to_write.setdefault(ri, {})[ci] = val

        src      = self._source
        token    = src["token"]
        sheet_id = src["sheet_id"]
        use_open = bool(src["use_open_api"])

        total = len(rows_to_write)
        self._status_var.set(f"正在写入 {total} 行到飞书...")
        self.update_idletasks()

        def worker():
            errors = []
            for ri, col_dict in rows_to_write.items():
                col_idxs = sorted(col_dict.keys())
                start_c  = col_idxs[0]
                end_c    = col_idxs[-1] + 1
                row_vals = [col_dict.get(ci, self._data[ri][ci] if ci < len(self._data[ri]) else "")
                            for ci in range(start_c, end_c)]
                try:
                    self.api.write_values(
                        token, sheet_id, [row_vals],
                        start_row=ri, start_col=start_c,
                        use_open_api=use_open,
                    )
                except Exception as ex:
                    errors.append(f"行 {ri}: {ex}")

            def done():
                if errors:
                    messagebox.showerror(
                        "部分写入失败",
                        f"成功 {total - len(errors)} 行，失败 {len(errors)} 行:\n"
                        + "\n".join(errors[:5]),
                        parent=self,
                    )
                else:
                    self._dirty.clear()
                    self._dirty_var.set("")
                    # 刷新 tag
                    for iid in self.tree.get_children():
                        ri2 = self._iid_to_row.get(iid)
                        if ri2 is not None:
                            tag = "even" if ri2 % 2 == 0 else "odd"
                            self.tree.item(iid, tags=(tag,))
                    messagebox.showinfo("保存成功",
                                        f"已写入 {total} 行到飞书", parent=self)
                self._status_var.set("保存完成")

            self.after(0, done)

        threading.Thread(target=worker, daemon=True).start()

    def _discard_dirty(self):
        if not self._dirty:
            return
        if not messagebox.askyesno("确认", f"撤销 {len(self._dirty)} 处修改并还原？",
                                   parent=self):
            return
        self._dirty.clear()
        self._dirty_var.set("")
        self._load_source()  # 重新从缓存加载

    # ── 粘贴一列 ─────────────────────────────────────────────────────────────

    def _paste_col_from_clipboard(self):
        """
        从剪贴板读取行数据（每行一个值，Tab/换行分隔的首列），
        写入当前表格的指定列，从指定起始行开始。
        常用场景：把 VLOOKUP 结果复制过来，一键填到目标列。
        """
        try:
            text = self.clipboard_get()
        except tk.TclError:
            messagebox.showwarning("提示", "剪贴板为空", parent=self)
            return

        try:
            col_idx  = int(self._paste_col_var.get().strip())
            start_ri = int(self._paste_row_var.get().strip())
        except ValueError:
            messagebox.showwarning("提示", "请输入有效的列号和起始行", parent=self)
            return

        lines  = text.strip().split("\n")
        values = [line.split("\t")[0].strip() for line in lines if line.strip()]

        if not values:
            messagebox.showinfo("提示", "剪贴板中没有可用数据", parent=self)
            return

        # 写入本地数据并标记脏
        changed = 0
        for offset, val in enumerate(values):
            ri = start_ri + offset
            while len(self._data) <= ri:
                self._data.append([""] * self._col_count)
            while len(self._data[ri]) <= col_idx:
                self._data[ri].append("")
            if self._data[ri][col_idx] != val:
                self._data[ri][col_idx] = val
                self._dirty[(ri, col_idx)] = val
                changed += 1

        # 刷新显示
        for iid in list(self.tree.get_children()):
            ri2 = self._iid_to_row.get(iid)
            if ri2 is not None and start_ri <= ri2 < start_ri + len(values):
                offset = ri2 - start_ri
                vals   = list(self.tree.item(iid, "values"))
                while len(vals) <= col_idx:
                    vals.append("")
                vals[col_idx] = values[offset]
                self.tree.item(iid, values=vals)
                self._refresh_row_tag(iid)

        self._dirty_var.set(f"● 未保存修改: {len(self._dirty)} 个单元格")
        self._status_var.set(
            f"已粘贴 {changed} 个值到列 {col_idx}，起始行 {start_ri}")


# ═════════════════════════════════════════════════════════════════════════════
# GUI 面板 5 — 使用示例
# ═════════════════════════════════════════════════════════════════════════════

class ExamplesPanel(tk.Frame):
    """每个函数的使用示例面板"""

    # ── 示例数据定义 ─────────────────────────────────────────────────────────

    EXAMPLES = [
        {
            "name": "VLOOKUP",
            "desc": "在「查找表」的指定列中搜索关键字，返回同一行的其他列数据。\n适合场景：用零件号查描述、价格、规格等。",
            "source_title": "查找表（零件信息表）",
            "source_cols": ["零件号", "描述", "单价", "库存"],
            "source_rows": [
                ["P001", "螺丝 M3×10", "0.05", "10000"],
                ["P002", "螺母 M3",    "0.03",  "8000"],
                ["P003", "垫片 M3",    "0.02",  "5000"],
                ["P004", "螺栓 M4×20", "0.12",  "3000"],
            ],
            "lookup_title": "查找值（需要查的零件号）",
            "lookup_items": ["P001", "P003", "P005（不存在）"],
            "steps": [
                ("查找值来源",   "从另一张表的「零件号」列，或直接输入"),
                ("查找范围",     "选「零件信息表」→ 查找列选「零件号」"),
                ("返回列",       "勾选「描述」「单价」（可多选）"),
                ("匹配模式",     "精确匹配"),
            ],
            "result_cols": ["查找值", "描述", "单价", "匹配状态"],
            "result_rows": [
                ["P001", "螺丝 M3×10", "0.05", "matched"],
                ["P003", "垫片 M3",    "0.02", "matched"],
                ["P005", "#N/A",       "#N/A", "not_found"],
            ],
        },
        {
            "name": "XLOOKUP",
            "desc": "搜索列和返回列可以来自【不同数据源】，突破 VLOOKUP 必须在同一张表的限制。\n适合场景：两张飞书表格之间做关联查找。",
            "source_title": "表A — 搜索列（供应商编码表）",
            "source_cols": ["供应商编码", "供应商名称"],
            "source_rows": [
                ["V001", "华勤科技"],
                ["V002", "富士康"],
                ["V003", "立讯精密"],
            ],
            "lookup_title": "表B — 返回列（价格表，与表A行序对应）",
            "lookup_items": ["表B：供应商编码 / 报价金额 / 交期(天)"],
            "steps": [
                ("查找值来源",   "你需要查找的供应商编码列表"),
                ("搜索表",       "选「供应商编码表」→ 搜索列选「供应商编码」"),
                ("返回表",       "选「价格表」→ 返回列选「报价金额」「交期」"),
                ("匹配模式",     "精确匹配"),
            ],
            "result_cols": ["查找值", "报价金额", "交期(天)", "匹配状态"],
            "result_rows": [
                ["V001", "¥12,500", "30", "matched"],
                ["V003", "¥9,800",  "45", "matched"],
                ["V999", "#N/A",    "#N/A", "not_found"],
            ],
        },
        {
            "name": "INDEX/MATCH",
            "desc": "MATCH 在一列中找到目标位置，INDEX 用该位置从另一列取值。\n两张表完全独立，最灵活，可跨任意表格任意列。",
            "source_title": "MATCH 表（用来定位行号）",
            "source_cols": ["物料编码", "物料描述"],
            "source_rows": [
                ["BOM-001", "主板组件"],
                ["BOM-002", "电源模块"],
                ["BOM-003", "散热风扇"],
            ],
            "lookup_title": "INDEX 表（从中取值，行序与 MATCH 表一致）",
            "lookup_items": ["INDEX 表：物料编码 / 成本 / 重量(g)"],
            "steps": [
                ("查找值",       "要查的物料编码"),
                ("MATCH 表",     "选「物料描述表」→ MATCH 列选「物料编码」"),
                ("INDEX 表",     "选「成本表」→ INDEX 列选「成本」「重量」"),
                ("匹配模式",     "精确 / 包含 均可"),
            ],
            "result_cols": ["查找值", "成本", "重量(g)", "匹配状态"],
            "result_rows": [
                ["BOM-001", "¥85.00", "320", "matched"],
                ["BOM-003", "¥12.50",  "85", "matched"],
                ["BOM-999", "#N/A",   "#N/A", "not_found"],
            ],
        },
        {
            "name": "SUMIF",
            "desc": "对满足条件的行求某列的数值之和，同时展示哪些行被纳入了计算。\n适合场景：按供应商、按品类统计金额。",
            "source_title": "数据表（采购记录）",
            "source_cols": ["供应商", "品类", "金额"],
            "source_rows": [
                ["华勤",   "电子", "12000"],
                ["富士康", "电子", "8500"],
                ["华勤",   "机械", "3200"],
                ["立讯",   "电子", "15000"],
                ["华勤",   "电子", "6800"],
            ],
            "lookup_title": "参数配置",
            "lookup_items": ["条件列：供应商", "条件值：华勤", "求和列：金额"],
            "steps": [
                ("数据源",   "选「采购记录表」"),
                ("条件列",   "供应商"),
                ("条件值",   "华勤（支持精确/包含/正则）"),
                ("求和列",   "金额"),
            ],
            "result_cols": ["条件值", "求和结果", "匹配行数"],
            "result_rows": [
                ["华勤", "22000", "3（行1、行3、行5）"],
            ],
        },
        {
            "name": "COUNTIF",
            "desc": "统计满足条件的行数，并展示哪些行被命中。\n适合场景：统计某供应商有多少条记录、某状态出现几次。",
            "source_title": "数据表（订单状态表）",
            "source_cols": ["订单号", "供应商", "状态"],
            "source_rows": [
                ["ORD-001", "华勤",   "已到货"],
                ["ORD-002", "富士康", "在途"],
                ["ORD-003", "华勤",   "已到货"],
                ["ORD-004", "华勤",   "缺货"],
                ["ORD-005", "立讯",   "已到货"],
            ],
            "lookup_title": "参数配置",
            "lookup_items": ["条件列：供应商", "条件值：华勤"],
            "steps": [
                ("数据源",   "选「订单状态表」"),
                ("条件列",   "供应商"),
                ("条件值",   "华勤"),
            ],
            "result_cols": ["条件值", "计数", "命中行"],
            "result_rows": [
                ["华勤", "3", "ORD-001、ORD-003、ORD-004"],
            ],
        },
        {
            "name": "SUMIFS",
            "desc": "多个条件同时满足时才求和，最多支持 4 个条件。\n适合场景：按供应商 + 品类 + 月份等多维度汇总。",
            "source_title": "数据表（采购记录，含多列条件）",
            "source_cols": ["供应商", "品类", "月份", "金额"],
            "source_rows": [
                ["华勤",   "电子", "1月", "12000"],
                ["富士康", "电子", "1月",  "8500"],
                ["华勤",   "机械", "1月",  "3200"],
                ["华勤",   "电子", "2月",  "6800"],
                ["华勤",   "电子", "1月",  "4500"],
            ],
            "lookup_title": "参数配置（2 个条件）",
            "lookup_items": [
                "条件1：供应商 = 华勤",
                "条件2：品类 = 电子",
                "条件3：月份 = 1月",
                "求和列：金额",
            ],
            "steps": [
                ("数据源",   "选「采购记录表」"),
                ("求和列",   "金额"),
                ("条件 1",   "供应商 = 华勤（精确匹配）"),
                ("条件 2",   "品类 = 电子（精确匹配）"),
                ("条件 3",   "月份 = 1月（精确匹配）"),
            ],
            "result_cols": ["条件组合", "求和结果", "命中行数"],
            "result_rows": [
                ["华勤 + 电子 + 1月", "16500", "2（行1、行5）"],
            ],
        },
    ]

    def __init__(self, master, **kwargs):
        super().__init__(master, bg=COLORS["bg"], **kwargs)
        self._build()

    def _build(self):
        tk.Label(self, text="使用示例",
                 bg=COLORS["bg"], fg=COLORS["text"],
                 font=("Microsoft YaHei UI", 13, "bold")).pack(
                     anchor="w", padx=16, pady=(14, 2))
        tk.Label(self,
                 text="每个函数的典型使用场景、参数配置说明和预期结果展示",
                 bg=COLORS["bg"], fg=COLORS["text_sub"],
                 font=("Microsoft YaHei UI", 9)).pack(anchor="w", padx=16, pady=(0, 10))

        nb = ttk.Notebook(self)
        nb.pack(fill="both", expand=True, padx=12, pady=(0, 12))

        for ex in self.EXAMPLES:
            page = self._make_example_page(nb, ex)
            nb.add(page, text=f"  {ex['name']}  ")

    def _make_example_page(self, parent, ex: dict) -> tk.Frame:
        outer = tk.Frame(parent, bg=COLORS["bg"])

        canvas = tk.Canvas(outer, bg=COLORS["bg"], highlightthickness=0)
        vsb = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        inner = tk.Frame(canvas, bg=COLORS["bg"])
        win_id = canvas.create_window((0, 0), window=inner, anchor="nw")

        def _on_configure(e):
            bbox = canvas.bbox("all")
            if bbox:
                canvas.configure(scrollregion=bbox)
        def _on_canvas_resize(e):
            canvas.itemconfig(win_id, width=e.width)

        def _bind_scroll(e):
            canvas.bind_all("<MouseWheel>",
                            lambda ev: canvas.yview_scroll(-1*(ev.delta//120), "units"))
        def _unbind_scroll(e):
            canvas.unbind_all("<MouseWheel>")

        inner.bind("<Configure>", _on_configure)
        canvas.bind("<Configure>", _on_canvas_resize)
        canvas.bind("<Enter>", _bind_scroll)
        canvas.bind("<Leave>", _unbind_scroll)

        pad = {"padx": 16, "pady": 6}

        # ── 函数描述 ─────────────────────────────────────────────────────────
        desc_card = make_card(inner)
        desc_card.pack(fill="x", **pad)
        tk.Label(desc_card, text=ex["name"], bg=COLORS["card"],
                 fg=COLORS["primary"], font=("Microsoft YaHei UI", 12, "bold")
                 ).pack(anchor="w", padx=12, pady=(10, 2))
        tk.Label(desc_card, text=ex["desc"], bg=COLORS["card"],
                 fg=COLORS["text"], font=("Microsoft YaHei UI", 9),
                 justify="left", wraplength=680
                 ).pack(anchor="w", padx=12, pady=(0, 10))

        # ── 示例数据 ─────────────────────────────────────────────────────────
        data_card = make_card(inner)
        data_card.pack(fill="x", **pad)
        tk.Label(data_card, text=ex["source_title"], bg=COLORS["card"],
                 fg=COLORS["text"], font=("Microsoft YaHei UI", 9, "bold")
                 ).pack(anchor="w", padx=12, pady=(8, 4))

        tree_frame = tk.Frame(data_card, bg=COLORS["card"])
        tree_frame.pack(fill="x", padx=12, pady=(0, 4))
        cols = ex["source_cols"]
        tree = ttk.Treeview(tree_frame, columns=cols, show="headings",
                            height=min(len(ex["source_rows"]), 5))
        for c in cols:
            tree.heading(c, text=c)
            tree.column(c, width=max(80, len(c)*14), anchor="center")
        for row in ex["source_rows"]:
            tree.insert("", "end", values=row)
        tree.pack(fill="x")

        # 查找值 / 额外说明
        if ex.get("lookup_items"):
            tk.Label(data_card, text=ex["lookup_title"], bg=COLORS["card"],
                     fg=COLORS["text"], font=("Microsoft YaHei UI", 9, "bold")
                     ).pack(anchor="w", padx=12, pady=(6, 2))
            for item in ex["lookup_items"]:
                tk.Label(data_card, text=f"  • {item}", bg=COLORS["card"],
                         fg=COLORS["text_sub"], font=("Microsoft YaHei UI", 9)
                         ).pack(anchor="w", padx=12)
        tk.Frame(data_card, height=8, bg=COLORS["card"]).pack()

        # ── 参数配置步骤 ─────────────────────────────────────────────────────
        step_card = make_card(inner)
        step_card.pack(fill="x", **pad)
        tk.Label(step_card, text="在「函数构建器」中这样配置",
                 bg=COLORS["card"], fg=COLORS["text"],
                 font=("Microsoft YaHei UI", 9, "bold")
                 ).pack(anchor="w", padx=12, pady=(8, 4))
        for i, (label, detail) in enumerate(ex["steps"], 1):
            row_f = tk.Frame(step_card, bg=COLORS["card"])
            row_f.pack(fill="x", padx=12, pady=2)
            tk.Label(row_f, text=f"  {i}.", bg=COLORS["card"],
                     fg=COLORS["primary"], font=("Microsoft YaHei UI", 9, "bold"),
                     width=3).pack(side="left")
            tk.Label(row_f, text=label, bg=COLORS["card"],
                     fg=COLORS["text"], font=("Microsoft YaHei UI", 9, "bold"),
                     width=12, anchor="w").pack(side="left")
            tk.Label(row_f, text=detail, bg=COLORS["card"],
                     fg=COLORS["text_sub"], font=("Microsoft YaHei UI", 9),
                     anchor="w").pack(side="left", fill="x", expand=True)
        tk.Frame(step_card, height=8, bg=COLORS["card"]).pack()

        # ── 预期结果 ─────────────────────────────────────────────────────────
        res_card = make_card(inner)
        res_card.pack(fill="x", **pad)
        tk.Label(res_card, text="预期执行结果", bg=COLORS["card"],
                 fg=COLORS["text"], font=("Microsoft YaHei UI", 9, "bold")
                 ).pack(anchor="w", padx=12, pady=(8, 4))
        res_frame = tk.Frame(res_card, bg=COLORS["card"])
        res_frame.pack(fill="x", padx=12, pady=(0, 4))
        rcols = ex["result_cols"]
        rtree = ttk.Treeview(res_frame, columns=rcols, show="headings",
                             height=len(ex["result_rows"]))
        for c in rcols:
            rtree.heading(c, text=c)
            rtree.column(c, width=max(90, len(c)*14), anchor="center")
        rtree.tag_configure("matched",   background="#e8f5e9", foreground="#2e7d32")
        rtree.tag_configure("not_found", background="#ffebee", foreground="#c62828")
        for row in ex["result_rows"]:
            tag = row[-1] if row[-1] in ("matched", "not_found") else ""
            display = row[:-1] if tag else row
            rtree.insert("", "end", values=display, tags=(tag,))
        rtree.pack(fill="x")
        tk.Frame(res_card, height=8, bg=COLORS["card"]).pack()

        return outer

# ═════════════════════════════════════════════════════════════════════════════
# GUI 面板 6 — API 设置
# ═════════════════════════════════════════════════════════════════════════════

class SettingsPanel(tk.Frame):

    def __init__(self, master, cache: DataCache, api: FeishuAPIClient, **kwargs):
        super().__init__(master, bg=COLORS["bg"], **kwargs)
        self.cache = cache
        self.api   = api
        self._build_ui()
        self._load_config()

    def _build_ui(self):
        tk.Label(self, text="API 设置", bg=COLORS["bg"],
                 fg=COLORS["text"], font=("Microsoft YaHei UI", 13, "bold")).pack(
                     anchor="w", padx=16, pady=(12, 4))

        card = make_card(self)
        card.pack(fill="x", padx=16, pady=8)

        fields = [
            ("内网代理地址", "proxy_url",
             f"默认: {FeishuAPIClient.DEFAULT_PROXY}"),
            ("App ID",       "app_id",
             f"默认: {FeishuAPIClient.DEFAULT_APP_ID}"),
            ("User ID",      "user_id",
             "企业内网用户 ID（工号）"),
            ("Access Token", "access_token",
             "可选 — 仅勾选「飞书开放平台 API」时需要填写，使用内网代理可留空"),
        ]

        self._vars: dict[str, tk.StringVar] = {}
        for ri, (label, key, hint) in enumerate(fields):
            is_token = key == "access_token"
            label_fg = COLORS["text_sub"] if is_token else COLORS["text"]
            tk.Label(card, text=label, bg=COLORS["card"], fg=label_fg,
                     font=("Microsoft YaHei UI", 9, "bold"),
                     width=14, anchor="e").grid(row=ri*2, column=0, padx=(12, 4), pady=(10, 0))
            var = tk.StringVar()
            self._vars[key] = var
            show = "*" if is_token else ""
            entry = ttk.Entry(card, textvariable=var, width=56, show=show)
            entry.grid(row=ri*2, column=1, padx=4, pady=(10, 0), sticky="ew")
            tk.Label(card, text=hint, bg=COLORS["card"], fg=COLORS["text_sub"],
                     font=("Microsoft YaHei UI", 7)).grid(
                         row=ri*2+1, column=1, padx=4, sticky="w")
        card.columnconfigure(1, weight=1)

        tk.Label(self, text="关于", bg=COLORS["bg"],
                 fg=COLORS["text"], font=("Microsoft YaHei UI", 10, "bold")).pack(
                     anchor="w", padx=16, pady=(16, 4))
        about_card = make_card(self)
        about_card.pack(fill="x", padx=16, pady=4)
        tk.Label(about_card,
                 text=f"{APP_NAME}  v{APP_VERSION}\n\n"
                      "基于飞书表格 API 的可视化函数面板\n"
                      "支持 VLOOKUP / XLOOKUP / INDEX-MATCH / SUMIF / COUNTIF / SUMIFS\n"
                      "数据本地缓存于 feishu_vlookup_data/cache.db",
                 bg=COLORS["card"], fg=COLORS["text_sub"],
                 font=("Microsoft YaHei UI", 9), justify="left").pack(
                     padx=16, pady=12)

        btn_row = tk.Frame(self, bg=COLORS["bg"])
        btn_row.pack(pady=8)
        make_label_btn(btn_row, "保存设置", self._save_config).pack(side="left", padx=6)
        make_label_btn(btn_row, "恢复默认", self._reset_defaults,
                       bg=COLORS["text_sub"]).pack(side="left", padx=6)

    def _load_config(self):
        cfg = self.cache.load_api_config()
        self._vars["proxy_url"].set(    cfg.get("proxy_url",     FeishuAPIClient.DEFAULT_PROXY))
        self._vars["app_id"].set(       cfg.get("app_id",        FeishuAPIClient.DEFAULT_APP_ID))
        self._vars["user_id"].set(      cfg.get("user_id",       ""))
        self._vars["access_token"].set( cfg.get("access_token",  ""))
        self._apply_to_api()

    def _save_config(self):
        self.cache.save_api_config({k: v.get() for k, v in self._vars.items()})
        self._apply_to_api()
        messagebox.showinfo("已保存", "设置已保存", parent=self)

    def _reset_defaults(self):
        self._vars["proxy_url"].set(FeishuAPIClient.DEFAULT_PROXY)
        self._vars["app_id"].set(FeishuAPIClient.DEFAULT_APP_ID)
        self._vars["user_id"].set("")
        self._vars["access_token"].set("")
        self._apply_to_api()

    def _apply_to_api(self):
        self.api.base_url      = self._vars["proxy_url"].get()    or FeishuAPIClient.DEFAULT_PROXY
        self.api.app_id        = self._vars["app_id"].get()       or FeishuAPIClient.DEFAULT_APP_ID
        self.api.user_id       = self._vars["user_id"].get()
        self.api.access_token  = self._vars["access_token"].get()


# ═════════════════════════════════════════════════════════════════════════════
# 主应用
# ═════════════════════════════════════════════════════════════════════════════

class FeishuFormulaApp(tk.Tk):

    def __init__(self):
        super().__init__()
        self.title(f"{APP_NAME}  v{APP_VERSION}")
        self.geometry("1180x820")
        self.minsize(900, 640)
        self.configure(bg=COLORS["bg"])

        self._setup_styles()

        self.cache = DataCache()
        self.api   = FeishuAPIClient()

        self._build_ui()
        self._load_api_config()

    def _setup_styles(self):
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("TNotebook",          background=COLORS["bg"],       borderwidth=0)
        style.configure("TNotebook.Tab",       background=COLORS["border"],
                        foreground=COLORS["text"], padding=(14, 6),
                        font=("Microsoft YaHei UI", 9))
        style.map("TNotebook.Tab",
                  background=[("selected", COLORS["primary"])],
                  foreground=[("selected", "white")])
        style.configure("Treeview",            background=COLORS["card"],
                        fieldbackground=COLORS["card"],
                        foreground=COLORS["text"], rowheight=24,
                        font=("Microsoft YaHei UI", 8))
        style.configure("Treeview.Heading",    background=COLORS["bg"],
                        foreground=COLORS["text_sub"],
                        font=("Microsoft YaHei UI", 8, "bold"))
        style.map("Treeview", background=[("selected", COLORS["highlight"])],
                  foreground=[("selected", COLORS["primary"])])
        style.configure("TEntry",             padding=4,
                        font=("Microsoft YaHei UI", 9))
        style.configure("TCombobox",          padding=4,
                        font=("Microsoft YaHei UI", 9))
        style.configure("TScrollbar",         background=COLORS["border"],
                        troughcolor=COLORS["bg"], width=8)

    def _build_ui(self):
        # ── 标题栏 ────────────────────────────────────────────────────────────
        header = tk.Frame(self, bg=COLORS["primary"], height=50)
        header.pack(fill="x")
        header.pack_propagate(False)
        tk.Label(header, text=f"  {APP_NAME}", bg=COLORS["primary"],
                 fg="white", font=("Microsoft YaHei UI", 12, "bold")).pack(
                     side="left", padx=8, pady=12)
        tk.Label(header, text=f"v{APP_VERSION}", bg=COLORS["primary"],
                 fg="#B0C4FF", font=("Microsoft YaHei UI", 8)).pack(side="left")

        # 快速说明
        quick_tips = (
            "① 数据源  ② 函数构建  ③ 查看结果 / 写回飞书  ④ 在线编辑表格  ⑤ API 设置"
        )
        tk.Label(header, text=quick_tips, bg=COLORS["primary"],
                 fg="#B0C4FF", font=("Microsoft YaHei UI", 8)).pack(side="right", padx=16)

        # ── 选项卡 ────────────────────────────────────────────────────────────
        self.nb = ttk.Notebook(self)
        self.nb.pack(fill="both", expand=True, padx=0, pady=0)

        self.results_panel = ResultsPanel(self.nb, self.cache, self.api)
        self.builder_panel = FunctionBuilderPanel(
            self.nb, self.cache,
            on_results=self._on_results
        )
        self.source_panel = DataSourcePanel(
            self.nb, self.cache, self.api,
            on_sources_changed=self._on_sources_changed
        )
        self.editor_panel   = OnlineEditorPanel(self.nb, self.cache, self.api)
        self.examples_panel = ExamplesPanel(self.nb)
        self.settings_panel = SettingsPanel(self.nb, self.cache, self.api)

        self.nb.add(self.source_panel,   text="  数据源管理  ")
        self.nb.add(self.builder_panel,  text="  函数构建器  ")
        self.nb.add(self.results_panel,  text="  执行结果    ")
        self.nb.add(self.editor_panel,   text="  在线编辑    ")
        self.nb.add(self.examples_panel, text="  使用示例    ")
        self.nb.add(self.settings_panel, text="  API 设置    ")

    def _on_results(self, headers, rows, func_type):
        # 推断 status 列索引
        status_col = None
        for ci, h in enumerate(headers):
            if h in ("匹配状态", "__status__"):
                status_col = ci
                break
        self.results_panel.show_results(headers, rows, func_type, status_col)
        self.nb.select(2)  # 自动跳到结果页

    def _on_sources_changed(self):
        self.builder_panel.reload_sources()
        self.editor_panel._reload_source_list()

    def _load_api_config(self):
        cfg = self.cache.load_api_config()
        if cfg:
            self.api.base_url     = cfg.get("proxy_url",    "") or FeishuAPIClient.DEFAULT_PROXY
            self.api.app_id       = cfg.get("app_id",       "") or FeishuAPIClient.DEFAULT_APP_ID
            self.api.user_id      = cfg.get("user_id",      "")
            self.api.access_token = cfg.get("access_token", "")


# ─── 入口 ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    app = FeishuFormulaApp()
    app.mainloop()
